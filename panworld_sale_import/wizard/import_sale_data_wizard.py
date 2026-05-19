# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) Fidobe Solutions LLC.
#
#    For Module Support : dev3@fidobe.com
#
#############################################################################
import binascii
import logging
import tempfile
import os
from pathlib import Path
import xlsxwriter

from odoo import _, fields, models
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger(__name__)


try:
    import xlrd
except ImportError:
    _logger.debug("Cannot `import xlrd`.")

try:
    import openpyxl
except ImportError:
    _logger.debug("Cannot `import openpyxl`.")
try:
    import base64
except ImportError:
    _logger.debug("Cannot `import base64`.")


class ImportSaleDataWizard(models.TransientModel):
    _name = "import.sale.data.wizard"
    _description = "Import Sale Data Wizard"

    data_import_file = fields.Binary(string="Select File")
    sample_file = fields.Binary("Download File")
    file_name = fields.Char(string="File Name")
    model_id = fields.Integer(string="Model Id")
    model_name = fields.Char(string="Model Name")

    def download_sample_file(self):
        """Method to download sample file for sale line import"""
        name_of_file = "sample_so_line_import.xls"
        fp = tempfile.NamedTemporaryFile(delete=False, suffix=".xls")
        path_list = list(os.path.split(fp.name))
        path_list[len(path_list) - 1] = name_of_file
        file_path = os.path.join(*path_list)
        workbook = xlsxwriter.Workbook(file_path)
        worksheet = workbook.add_worksheet()
        counter = 0

        context = self.env.context
        model_name = self.model_name
        model_id = self.model_id
        model_rec = self.env[model_name].browse([model_id])

        field_list_to_import = [
            "id",
            "isbn",
            "word_count",
            "cost_per_unit",
            "quotation qty",
            "Quote Cancel Qty",
            # "so quantity",
            "cancelled qty",
            "Quote Cancel Reason",
            "cancel reason",
            "unit price",
            "discount",
            "taxes",
            "Grade",
            "Subject",
            "Format",
            "Classification",
            "remarks",
            "course id",
            "crn",
            "csn_no",
        ]
        row = 1
        for line in model_rec.order_line:
            quote_cancel_reason_selection = [l[0] for l in line._fields['quote_cancel_reason'].selection]
            worksheet.data_validation(row, 5, row, 5, {'validate': 'list', 'source': quote_cancel_reason_selection})
            row += 1
            cancel_reason_selection = [l[0] for l in line._fields['cancel_reason'].selection]
            worksheet.data_validation(row, 6, row, 6, {'validate': 'list', 'source': cancel_reason_selection})
            row += 1
        for i in field_list_to_import:
            worksheet.write(0, counter, i)
            counter += 1
        workbook.close()
        export_id = base64.b64encode(open(file_path, "rb+").read())
        self.write({"sample_file": export_id, "file_name": name_of_file})
        return {
            "name": "Import Order Lines",
            "view_mode": "form",
            "res_id": self.id,
            "res_model": "import.sale.data.wizard",
            "view_type": "form",
            "type": "ir.actions.act_window",
            "target": "new",
        }

    def import_data(self):
        """Method to import sale order lines"""
        if not self.data_import_file:
            raise UserError(_("Please upload file!."))
        context = self.env.context
        model_name = self.model_name
        model_id = self.model_id
        model_rec = self.env[model_name].browse([model_id])
        fp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        fp.write(binascii.a2b_base64(self.data_import_file))
        fp.seek(0)
        try:
            # Try opening as xlsx with openpyxl
            workbook = openpyxl.load_workbook(fp.name, data_only=True, read_only=True)
            sheet = workbook.active
            headers = []
            for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = list(row)
                break
            
            codes = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                 if len(row) > 1 and row[1]:
                     codes.append(str(row[1]).split('.')[0])
            is_xlsx = True
        except Exception as e:
            _logger.warning("Failed to open with openpyxl: %s", e)
            # Fallback to xlrd for xls
            is_xlsx = False
            sheet = xlrd.open_workbook(fp.name).sheet_by_index(0)
            headers = []
            codes = []
            for i in range(sheet.ncols):
                headers.append(sheet.cell_value(0, i))
            for rownum in range(1, sheet.nrows):
                codes.append(str(sheet.cell_value(rownum, 1)).split('.')[0])
                
        # codes.pop(0) # Removed because we now start iteration from data rows directly
        msg = ""
        for code in codes:
            product = self.env["product.product"].search([("barcode", "=", code)])
            if not product:
                msg += (f"[{code}] ")
        if msg:
            raise UserError(_("No product(s) Found with ISBN : \n" + msg))
        file_headers = [str(x or '').lower() for x in headers]
        res = False
        list_records = []
        list_crn = []
        without_crn_list = []
        field_list_to_import = [
            "id",
            "isbn",
            "word_count",
            "cost_per_unit",
            "quotation qty",
            "quote cancel qty",
            # "so quantity",
            "cancelled qty",
            "quote cancel reason",
            "cancel reason",
            "unit price",
            "discount",
            "taxes",
            "grade",
            "subject",
            "format",
            "classification",
            "remarks",
            "course id",
            "crn",
            "csn_no",
        ]
        file_headers = [h for h in file_headers if h]
        if file_headers != field_list_to_import:
            raise UserError(
                _("File headers must match the sample file's headers. \nExpected: %s \nFound: %s") % (field_list_to_import, file_headers))
        # if self._context.get('active_model') != "sale.order":
        #     raise UserError(_('You can not perform both operation at a time. Please close the wizard and try to Import the file again.'))
        for line in model_rec.order_line:
            if line.product_id.barcode in codes and model_rec.state in ("sent", "draft"):
                continue
            elif line.product_id.barcode not in codes and model_rec.state in ("sent", "draft"):
                line.unlink()
        vals_list = []
        
        if is_xlsx:
             workbook = openpyxl.load_workbook(fp.name, data_only=True, read_only=True)
             sheet = workbook.active
             for row in sheet.iter_rows(min_row=2, values_only=True):
                 passing_dict = dict(zip(field_list_to_import, row))
                 if not passing_dict.get('isbn'):
                     continue
                 self.create_update_line(passing_dict, model_rec, model_name, vals_list)

        else:
             # xlrd iteration
             for row_no in range(sheet.nrows):
                if row_no != 0:
                    passing_dict = dict(
                        zip(field_list_to_import, sheet.row_values(row_no)))
                    if not passing_dict.get('isbn'):
                        continue
                    self.create_update_line(passing_dict, model_rec, model_name, vals_list)
        if model_name == 'sale.order':
            self.env['sale.order.line'].create(vals_list)
            # for line in model_rec.order_line:
            #     line._onchange_uom_qty_discount_price_unit()
            #     line.onchange_product_uom_qty()
        return res

    def is_number(self, data):
        try:
            data = float(data)
            if isinstance(data, float):
                return True
        except ValueError:
            return False

    def validate_and_get_crn(self, value):
        """Method to get crn"""
        crn = value.get("crn")
        if not crn:
            return False
        if self.is_number(crn):
            return int(float(crn))
        return crn

    def validate_and_get_course_id(self, values):
        """Method to get course_id"""
        course_id = values.get("course id")
        if not course_id:
            return ''
        if self.is_number(course_id):
            return int(float(course_id))
        return course_id

    def validate_and_get_item_code(self, value):
        """Method to get isbn(barcode) code of product"""
        item_code = value.get("isbn")
        if not item_code:
            raise UserError(_("Isbn is required."))
        item_code = str(item_code).strip()
        if len(item_code) <= 0:
            raise UserError(_("Isbn is required."))
        if self.is_number(item_code):
            return int(float(item_code))
        return item_code

    # def validate_and_get_analytic_tag(self, value):
    #     """Method to get analytic tag"""
    #     analytic_tags = value.get("analytic_tags")
    #     if not analytic_tags:
    #         raise UserError(_("Analytic tags are required."))
    #     analytic_tags = str(analytic_tags).split(",")
    #     if len(analytic_tags) <= 0:
    #         raise UserError(_("Analytic tags are required."))
    #     return analytic_tags

    def validate_and_get_done_qty(self, value):
        """Method to update get done qty"""
        done = value.get("done")
        if not done:
            raise UserError(_("Done is required."))
        done = str(done).strip()
        if len(done) <= 0:
            raise UserError(_("Done is required."))
        if self.is_number(done):
            return int(float(done))
        return done

    def validate_and_get_tax(self, value):
        """Method to get Tax names as list"""
        tax_ids = value.get("taxes")
        if not tax_ids:
            return []
        # Split by comma and remove extra spaces
        tax_ids = [tax.strip() for tax in str(tax_ids).split(",")]
        return tax_ids

    def create_update_line(self, values, model_rec, model_name, vals_list=[]):
        """Method to update the done qty of picking order"""
        code = self.validate_and_get_item_code(values)
        # analytic_tags = self.validate_and_get_analytic_tag(values)
        # analytic_tags = False
        so_line_obj = self.env["sale.order.line"]
        product = self.env["product.product"].search([("barcode", "=", code),('active', '=', True)])
        # analytic_tag_obj = self.env["account.analytic.tag"]
        grade = values.get("grade")
        product_grade = self.env['product.grade'].search([('name', '=', grade)])
        if grade and not product_grade:
            raise ValidationError(_(f"No Grade named {grade} can be found"))
        subject = values.get("subject")
        product_subject = self.env['product.subject'].search([('name', '=', subject)])
        if subject and not product_subject:
            raise ValidationError(_(f"No Subject named {subject} can be found"))
        tax_list = []
        if not product:
            raise UserError(_(f"No product Found with isbn: {code}"))
        analytic_tag_ids = []
        # for analytic_tag in analytic_tags:
        #     analytic_tag = analytic_tag.strip()
        #     analytic_tag_rec = analytic_tag_obj.search(
        #         [("name", "=", analytic_tag)], limit=1
        #     )
        #     if not analytic_tag_rec:
        #         raise UserError(
        #             _("Analytic tags %s not found in system!", analytic_tag)
        #         )
        #     analytic_tag_ids.append(analytic_tag_rec.id)
        tax_names = self.validate_and_get_tax(values)
        if tax_names:
            domain = [("name", "in", tax_names), ("type_tax_use", "=", 'sale'), ("company_id", "in", [False, model_rec.company_id.id])]
            taxes = self.env["account.tax"].search(domain)
            found_tax_names = set(taxes.mapped('name'))
            missing_taxes = set(tax_names) - found_tax_names
            if missing_taxes:
                msg = ""
                for missing_tax in missing_taxes:
                    # Try to find the tax in other companies to use as template
                    # Use sudo to ensure we can see taxes from other companies
                    alt_taxes = self.env["account.tax"].sudo().search([("name", "=", missing_tax), ("type_tax_use", "=", 'sale')], limit=1)
                    
                    if alt_taxes:
                        # Auto-create the tax for the current company
                        new_tax_vals = {
                            'name': alt_taxes.name,
                            'amount': alt_taxes.amount,
                            'amount_type': alt_taxes.amount_type,
                            'type_tax_use': alt_taxes.type_tax_use,
                            'company_id': model_rec.company_id.id,
                            'description': alt_taxes.description,
                            'price_include': alt_taxes.price_include,
                            'include_base_amount': alt_taxes.include_base_amount,
                        }
                        try:
                            new_tax = self.env['account.tax'].create(new_tax_vals)
                            taxes += new_tax
                            _logger.info(f"Auto-created missing tax '{missing_tax}' for company '{model_rec.company_id.name}' based on template from '{alt_taxes.company_id.name}'.")
                        except Exception as e:
                            msg += f"\n- Failed to auto-create '{missing_tax}': {str(e)}"
                    else:
                        msg += f"\n- '{missing_tax}' not found in the system."
                
                if msg:
                     raise UserError(_("Tax Validation Error:%s") % msg)
            
            tax_list = [(6, 0, taxes.ids)]

        if model_name == "sale.order":
            so_qty = float(values.get("quotation qty")) if values.get("quotation qty") else 0
            quote_cancel_qty = float(values.get("quote cancel qty")) if values.get("quote cancel qty") else 0
            so_quantity = float(values.get("so quantity")) if values.get("so quantity") else 0
            cancelled_qty = float(values.get("cancelled qty")) if values.get("cancelled qty") else 0
            quote_cancel_reason = values.get("quote cancel reason")
            cancel_reason = values.get("cancel reason")
            price = values.get("unit price")
            discount = values.get("discount")
            remarks = values.get("remarks")
            csn_no = values.get("csn_no")
            word_count = values.get("word_count")
            cost_per_unit = values.get("cost_per_unit")

            if word_count and cost_per_unit:
                price = word_count * cost_per_unit
                so_qty = 1.0
                so_quantity = 1.0
            so_line_id = int(values.get("id")) if values.get("id") else False
            grade = product_grade.id if product_grade else False
            subject = product_subject.id if product_subject else False
            format = values.get("format")
            classification = values.get("classification")
            course_id = self.validate_and_get_course_id(values)
            crn = self.validate_and_get_crn(values)
            # as per client requirements bypass warning base on customer vat number
            if model_rec and model_rec.partner_id.vat != '100276002100003':
                # so_line_rec = so_line_obj.search(
                #     [
                #         ("product_id.barcode", "=", code),
                #         ("order_id", "=", model_rec and model_rec.id),
                #     ]
                # )
                so_line_rec = so_line_obj.browse(so_line_id) if so_line_id else so_line_obj
                vals = {
                    "so_qty": so_qty,
                    "quote_cancel_qty": quote_cancel_qty,
                    "so_quantity": so_qty,
                    "cancelled_qty": cancelled_qty,
                    "quote_cancel_reason": quote_cancel_reason,
                    "cancel_reason": cancel_reason,
                    "price_unit": price,
                    "discount": discount,
                    # "analytic_tag_ids": [(6, 0, analytic_tag_ids)],
                    "remarks": remarks,
                    "course_id": course_id,
                    "crn": crn,
                    "csn_no": csn_no,
                    "word_count": word_count,
                    "cost_per_unit": cost_per_unit,
                    "uk_wholesaler_id":grade,
                    "subject_id":subject,
                    "format":format,
                    "classification":classification,
                    "tax_ids":tax_list
                }
                if model_rec and model_rec.state in ("sent", "draft"):
                    if vals.get('quote_cancel_qty') > 0 and not vals.get('quote_cancel_reason'):
                        # title = _("Quote Cancel Reason Required!")

                        # message = _(
                        #     "Data Not Imported! Please check your import file when you add quote cancel qty, quote cancel reason is required.")

                        raise UserError(_('Data Not Imported! Please check your import file when you add quote cancel qty, quote cancel reason is required.'))
                        # return {
                        #     'type': 'ir.actions.client',
                        #     'tag': 'display_notification',
                        #     'params': {
                        #         'title': title,
                        #         'message': message,
                        #         # 'sticky': False,
                        #         'type': 'warning',
                        #     }
                        # }
                    if not so_line_rec:
                        vals.update({
                            "product_id": product and product.id,
                            "order_id": model_rec and model_rec.id,
                            "product_uom_qty": so_qty - quote_cancel_qty,
                            "so_quantity": so_qty - quote_cancel_qty,
                            "list_price": price - (price * discount / 100.0) if discount else price
                        })
                        vals_list.append(vals)
                    else:
                        vals.update({
                            "product_uom_qty": so_qty - quote_cancel_qty,
                            "so_quantity": so_qty - quote_cancel_qty,
                            "list_price": price - (price * discount / 100.0) if discount else price
                        })
                        so_line_rec.write(vals)
                elif model_rec and model_rec.state == "sale" and so_line_rec:
                    if vals.get('cancelled_qty') > 0 and not vals.get('cancel_reason'):
                        title = _("Cancel Reason Required!")
                        raise UserError(
                            _('Data Not Imported! Please check your import file when you add quote cancel qty, quote cancel reason is required.'))

                        # message = _(
                        #     "Data Not Imported! Please check your import file when you add cancelled qty, cancel reason is required.")
                        # return {
                        #     'type': 'ir.actions.client',
                        #     'tag': 'display_notification',
                        #     'params': {
                        #         'title': title,
                        #         'message': message,
                        #         # 'sticky': False,
                        #         'type': 'warning',
                        #     }
                        # }
                    vals.update({
                        # "product_uom_qty": so_qty - cancelled_qty,
                        "product_uom_qty": (so_qty - quote_cancel_qty) - cancelled_qty,
                        "so_quantity": so_qty - quote_cancel_qty,
                        "list_price": price - (price * discount / 100.0) if discount else price
                    })
                    so_line_rec.write(vals)
            else:
                # so_line_rec = so_line_obj.search(
                #     [
                #         ("product_id.barcode", "=", code),
                #         ("order_id", "=", model_rec and model_rec.id),
                #         ("crn", "=", crn),
                #     ]
                # )
                so_line_rec = so_line_obj.browse(so_line_id) if so_line_id else so_line_obj
                vals = {
                    "so_qty": so_qty,
                    "quote_cancel_qty": quote_cancel_qty,
                    "so_quantity": so_qty,
                    "cancelled_qty": cancelled_qty,
                    "quote_cancel_reason": quote_cancel_reason,
                    "cancel_reason": cancel_reason,
                    "price_unit": price,
                    "discount": discount,
                    # "analytic_tag_ids": [(6, 0, analytic_tag_ids)],
                    "remarks": remarks,
                    "format": format,
                    "classification": classification,
                    "course_id": course_id,
                    "crn": crn,
                    "csn_no": csn_no,
                    "word_count": word_count,
                    "cost_per_unit": cost_per_unit,
                }
                if model_rec and model_rec.state in ("sent", "draft"):
                    if vals.get('quote_cancel_qty') > 0 and not vals.get('quote_cancel_reason'):
                        # title = _("Quote Cancel Reason Required!")
                        raise UserError(
                            _('Data Not Imported! Please check your import file when you add quote cancel qty, quote cancel reason is required.'))

                        # message = _(
                        #     "Data Not Imported! Please check your import file when you add quote cancel qty, quote cancel reason is required.")
                        # return {
                        #     'type': 'ir.actions.client',
                        #     'tag': 'display_notification',
                        #     'params': {
                        #         'title': title,
                        #         'message': message,
                        #         # 'sticky': False,
                        #         'type': 'warning',
                        #     }
                        # }
                    if not so_line_rec:
                        vals.update({
                            "product_id": product and product.id,
                            "order_id": model_rec and model_rec.id,
                            "product_uom_qty": so_qty - quote_cancel_qty,
                            "so_quantity": so_qty - quote_cancel_qty,
                            "list_price": price - (price * discount / 100.0) if discount else price
                        })
                        vals_list.append(vals)
                    else:
                        vals.update({
                            "product_uom_qty": so_qty - quote_cancel_qty,
                            "so_quantity": so_qty - quote_cancel_qty,
                            "list_price": price - (price * discount / 100.0) if discount else price
                        })
                        so_line_rec.write(vals)
                elif model_rec and model_rec.state == "sale" and so_line_rec:
                    if vals.get('cancelled_qty') > 0 and not vals.get('cancel_reason'):
                        title = _("Cancel Reason Required!")
                        raise UserError(
                            _('Data Not Imported! Please check your import file when you add quote cancel qty, quote cancel reason is required.'))

                        # message = _(
                        #     "Data Not Imported! Please check your import file when you add cancelled qty, cancel reason is required.")
                        # return {
                        #     'type': 'ir.actions.client',
                        #     'tag': 'display_notification',
                        #     'params': {
                        #         'title': title,
                        #         'message': message,
                        #         # 'sticky': False,
                        #         'type': 'warning',
                        #     }
                        # }
                    vals.update({
                        "product_uom_qty": (so_qty - quote_cancel_qty) - cancelled_qty,
                        # "product_uom_qty": so_qty - cancelled_qty,
                        "so_quantity": so_qty - quote_cancel_qty,
                        "list_price": price - (price * discount / 100.0) if discount else price
                    })
                    so_line_rec.write(vals)
        return vals_list

    # def create_attachment(self, file_name):
    #     """
    #     Delete file created in tmp dir, Delete old attachment and create attachment for download report
    #     :param file_name: file name string
    #     :return: attachment ir.attachment object
    #     """
    #     ir_attachment_obj = self.env['ir.attachment']
    #     # Read File data
    #     with open(f'/tmp/{file_name}', "rb+") as file:
    #         file_data = base64.encodebytes(file.read())
    #         file.close()
    #
    #     # Remove tmp file
    #     os.remove(f'/tmp/{file_name}')
    #     # Delete Old Attachment
    #     attachments = ir_attachment_obj.search([('name', '=ilike', file_name),
    #                                             ('res_model', '=', 'purchase.order')])
    #     attachments and attachments.unlink()
    #
    #     return ir_attachment_obj.create({
    #         'name': file_name,
    #         'datas': file_data,
    #         'res_model': 'purchase.order',
    #         'type': 'binary'
    #     })

    def create_attachment(self, file_name):
        """
        Delete file created in tmp dir, Delete old attachment and create attachment for download report
        :param file_name: file name string
        :return: attachment ir.attachment object
        """
        ir_attachment_obj = self.env['ir.attachment']

        if not file_name or not isinstance(file_name, str):
            raise ValueError("Invalid file_name provided")

        file_name = Path(file_name).name
        if not file_name or file_name in ['.', '..', '']:
            raise ValueError("Invalid file_name after sanitization")

        file_path = f'/tmp/{file_name}'
        if os.path.exists(file_path):
            if os.path.isdir(file_path):
                raise IsADirectoryError(f"Path {file_path} is a directory, not a file")
        else:
            raise FileNotFoundError(f"File {file_path} does not exist")

        try:
            with open(file_path, "rb") as file:
                file_data = base64.encodebytes(file.read())
            os.remove(file_path)
            attachments = ir_attachment_obj.search([
                ('name', '=ilike', file_name),
                ('res_model', '=', 'purchase.order')
            ])
            if attachments:
                attachments.unlink()

            return ir_attachment_obj.create({
                'name': file_name,
                'datas': file_data,
                'res_model': 'purchase.order',
                'type': 'binary'
            })

        except Exception as e:
            raise

    def export_data(self):
        report_file_name = ''
        """Method to export purchase order lines with difference for compare actions"""
        if self.env.context.get('active_model') == 'sale.order':
            report_file_name = self.export_so_data()
        # Create Attachment
        attachment = self.create_attachment(report_file_name)
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'download',
        }

    def export_so_data(self):
        active_id = self.env.context.get("active_id")
        so_id = self.env["sale.order"].browse(active_id)
        so_lines = so_id.order_line
        file_name = 'so_lines_data.xlsx'
        workbook = xlsxwriter.Workbook(f'/tmp/{file_name}')
        worksheet = workbook.add_worksheet("POLinesData")
        worksheet.protect()
        worksheet.set_landscape()
        worksheet.fit_to_pages(1, 0)
        worksheet.set_zoom(80)
        worksheet.set_column(1, 1, 20)
        worksheet.set_column(2, 2, 20)
        worksheet.set_column(3, 3, 15)
        worksheet.set_column(4, 4, 15)
        worksheet.set_column(6, 6, 20)

        qty_data = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'locked': False})
        qty_data_lock = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
        header_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'bold': True, 'border': 1})

        row, col = 0, 0
        worksheet.set_row(1, 30)
        worksheet.write(row, col, 'id', header_format)
        col += 1
        worksheet.write(row, col, 'isbn', header_format)
        col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'word_count', header_format)
        col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'cost_per_unit', header_format)
        col += 1
        worksheet.write(row, col, 'quotation qty', header_format)
        col += 1
        worksheet.set_column(col, col, 20)
        worksheet.write(row, col, 'Quote Cancel Qty', header_format)
        col += 1
        # worksheet.write(row, col, 'So quantity', header_format)
        # col += 1
        worksheet.set_column(col, col, 40)
        worksheet.write(row, col, 'Cancelled Qty', header_format)
        col += 1
        worksheet.set_column(col, col, 20)
        worksheet.write(row, col, 'Quote Cancel Reason', header_format)
        col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'Cancel reason', header_format)
        col += 1
        worksheet.write(row, col, 'unit price', header_format)
        col += 1
        worksheet.write(row, col, 'discount', header_format)
        col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'taxes', header_format)
        col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'Grade', header_format)
        col +=1
        worksheet.write(row, col, 'Subject', header_format)
        col += 1
        worksheet.write(row, col, 'Format', header_format)
        col += 1
        worksheet.write(row, col, 'Classification', header_format)
        col += 1
        worksheet.set_column(col, col, 20)
        worksheet.write(row, col, 'remarks', header_format)
        col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'course id', header_format)
        col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'crn', header_format)
        col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'csn_no', header_format)
        col += 1

        row = 1

        for line in so_lines:
            col = 0
            worksheet.write(row, col, line.id or '', qty_data_lock)
            col += 1
            worksheet.write(row, col, line.product_id.barcode or '', qty_data_lock)
            col += 1
            worksheet.write(row, col, line.word_count or 0, qty_data_lock)
            col += 1
            worksheet.write(row, col, line.cost_per_unit or 0, qty_data_lock)
            col += 1
            worksheet.write(row, col, line.so_qty or 0, qty_data_lock)
            col += 1
            worksheet.write(row, col, line.quote_cancel_qty or 0, qty_data) if line.state in ['draft', 'sent'] else worksheet.write(row, col, line.quote_cancel_qty or 0, qty_data_lock)
            # col += 1
            # worksheet.write(row, col, line.so_quantity or 0, qty_data_lock)
            col += 1
            worksheet.write(row, col, line.cancelled_qty or 0, qty_data)
            col += 1
            quote_cancel_reason_selection = [l[0] for l in line._fields['quote_cancel_reason'].selection]
            worksheet.data_validation(row, 7, row, 7, {'validate': 'list', 'source': quote_cancel_reason_selection})
            worksheet.write(row, col, line.quote_cancel_reason or '', qty_data)
            col += 1
            cancel_reason_selection = [l[0] for l in line._fields['cancel_reason'].selection]
            worksheet.data_validation(row, 8, row, 8, {'validate': 'list', 'source': cancel_reason_selection})
            worksheet.write(row, col, line.cancel_reason or '', qty_data)
            col += 1
            worksheet.write(row, col, line.list_price or 0, qty_data)
            col += 1
            worksheet.write(row, col, line.discount or 0, qty_data)
            col += 1
            taxes = ', '.join(line.tax_ids.mapped('name'))
            worksheet.write(row, col, taxes or '', qty_data_lock)
            col += 1
            worksheet.write(row, col, line.uk_wholesaler_id.name or '', qty_data_lock)
            col += 1
            worksheet.write(row, col, line.subject_id.name or '', qty_data)
            col += 1
            worksheet.write(row, col, line.format or '', qty_data)
            col += 1
            worksheet.write(row, col, line.classification or '', qty_data)
            col += 1
            worksheet.write(row, col, line.remarks or '', qty_data)
            col += 1
            worksheet.write(row, col, line.course_id or '', qty_data_lock)
            col += 1
            worksheet.write(row, col, line.crn or '', qty_data_lock)
            col += 1
            worksheet.write(row, col, line.csn_no or '', qty_data_lock)
            col += 1
            row += 1
        workbook.close()
        return file_name
