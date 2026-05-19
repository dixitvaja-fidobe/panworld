# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) Fidobe Solutions LLC.
#
#    For Module Support : dev3@fidobe.com
#
#############################################################################
import base64
import binascii
import datetime
import os
import tempfile

import xlrd
import openpyxl
import xlsxwriter

from odoo import _, fields, models
from odoo.exceptions import UserError


class ImportProduct(models.TransientModel):
    _name = "import.product"
    _description = "Import Product"

    file = fields.Binary("File")
    sample_file = fields.Binary("Download File")
    file_name = fields.Char(string="File Name")



    field_list_to_import = [
        "Item Code",
        "Item Name",
        "Item Category",
        "Cost Price",
        "Weight In (KG)",
        "Selling Price",
        "Author",
        "Publisher",
        "Master Publisher",
        "Uk Wholesaler",
        "Non-Uk Wholesalers",
        "Publication Country",
        "Main Title",
        "Subtitle",
        "Subject",
        "Pages",
        "Edition",
        "Series",
        "Book Language",
        "Product Format",
        "Interest Age",
        "Audience Readership",
        "Publication Date",
        "Status",
        "Height Along Spine In Mm",
        "Width From Spine To Edge In Mm",
        "Product Type",
        "Can be sold",
        "Can be purchased",
        "Available in POS",
        "Customer Taxes",
        "Vendor Taxes"
    ]

    req_field_list_to_import = [
        "Item Code",
        "Item Name",
        "Item Category",
        "Weight In (KG)",
        "Product Format",
        "Product Type",
        "Customer Taxes",
        "Vendor Taxes"
    ]
    def download_sample_file(self):
        name_of_file = "sample_with_product.xls"
        fp = tempfile.NamedTemporaryFile(delete=False, suffix=".xls")
        path_list = list(os.path.split(fp.name))
        path_list[len(path_list) - 1] = name_of_file
        file_path = os.path.join(*path_list)
        workbook = xlsxwriter.Workbook(file_path)
        worksheet = workbook.add_worksheet()
        counter = 0
        header_format = workbook.add_format({
            'bg_color': 'yellow',
        })
        for i in self.field_list_to_import:
            if i not in self.req_field_list_to_import:
                worksheet.write(0, counter, i)
            else:
                worksheet.write(0, counter, i, header_format)
            counter += 1
        workbook.close()
        export_id = base64.b64encode(open(file_path, "rb+").read())
        self.write({"sample_file": export_id, "file_name": name_of_file})
        return {
            "name": "Import Products",
            "view_mode": "form",
            "res_id": self.id,
            "res_model": "import.product",
            "type": "ir.actions.act_window",
            "target": "new",
        }

    def get_or_create_author(self, value):
        author = value.get("Author")
        if author:
            author = author.strip()
            if len(author) > 0:
                author_id = self.env["res.partner"].search([("name", "=ilike", author)])
                if not author_id:
                    author_id = self.env["res.partner"].create({"name": author})
                return author_id.id

    def get_or_create_publisher(self, value):
        publisher = value.get("Publisher")
        if publisher:
            publisher = publisher.strip()
            if len(publisher) > 0:
                publisher_id = self.env["res.partner"].search(
                    [("name", "=ilike", publisher)], limit=1
                )
                if not publisher_id:
                    publisher_id = self.env["res.partner"].create({"name": publisher})
                return publisher_id.id

    def get_or_create_master_publisher(self, value):
        master_publisher = value.get("Master Publisher")
        if master_publisher:
            master_publisher = master_publisher.strip()
            if len(master_publisher) > 0:
                master_publisher_obj = self.env["res.partner"].search(
                    [("name", "=ilike", master_publisher)], limit=1
                )
                if not master_publisher_obj:
                    master_publisher_obj = self.env["res.partner"].create(
                        {"name": master_publisher}
                    )
                return master_publisher_obj.id

    def get_item_category(self, value):
        item_category = value.get("Item Category")
        if item_category:
            product_category = self.env["product.category"].search(
                [("name", "=ilike", item_category)], limit=1
            )
            if product_category:
                if len(product_category) > 1:
                    product_category = product_category[0]
                return product_category.id
            else:
                raise UserError(_(f"Item Category is not found: {item_category}"))
        else:
            raise UserError(_("Item Category is required."))

    def is_number(self, data):
        try:
            data = float(data)
            if isinstance(data, float):
                return True
        except ValueError:
            return False

    def validate_and_get_item_code(self, value):
        item_code = value.get("Item Code")
        if item_code:
            item_code = str(item_code)
            item_code = item_code.strip()
            if len(item_code) <= 0:
                raise UserError(_("Item code is required."))
            if self.is_number(item_code):
                return int(float(item_code))
            return item_code
        else:
            raise UserError(_("Item code is required."))

    def validate_and_get_item_name(self, value):
        item_name = value.get("Item Name")
        if item_name:
            if isinstance(item_name, float):
                item_name = int(item_name)
            if not isinstance(item_name, str):
                item_name = str(item_name)
            item_name = item_name.strip()
            if len(item_name) < 0:
                raise UserError(_("Item name is required."))
            return item_name
        else:
            raise UserError(_("Item name is required."))

    def validate_and_get_weight(self, value):
        detailed_type = value.get("Product Type")
        import_product = self.env.context.get("import_product")
        item_category = value.get("Item Category")
        weight_in_kg = value.get("Weight In (KG)")
        is_weight_required = True
        if detailed_type and detailed_type in ["Service", "Consumable"]:
            is_weight_required = False
        if not weight_in_kg and import_product and is_weight_required:
            raise UserError(_("Weight is required."))
        if type(weight_in_kg) is float:
            if weight_in_kg < 0:
                raise UserError(_("Weight value must be positive."))
            else:
                return float(weight_in_kg)
        return 0.0

    def get_publication_country(self, value):
        country_name = value.get("Publication Country")
        if country_name:
            country = self.env["res.country"].search([("name", "=", country_name)], limit=1)
            if country:
                return country.id
        return False

    def get_or_create_wholesaler(self, value, is_uk_wholesaler):
        model_name = "res.partner"
        if is_uk_wholesaler:
            wholesaler = value.get("Uk Wholesaler")
        else:
            wholesaler = value.get("Non-Uk Wholesalers")
        if wholesaler:
            wholesaler_obj = self.env[model_name].search(
                [("name", "=ilike", wholesaler)], limit=1
            )
            if not wholesaler_obj:
                wholesaler_obj = self.env[model_name].create({"name": wholesaler})
            return wholesaler_obj.id

    def get_cost_price(self, value):
        standard_price = value.get("Cost Price")
        if isinstance(standard_price, str):
            standard_price = standard_price.strip()
            standard_price = float(standard_price) if len(standard_price) > 0 else 0.0
        elif isinstance(standard_price, float):
            if standard_price <= 0.0:
                return 0.01
            return standard_price
        elif isinstance(standard_price, int):
            if standard_price <= 0:
                return 0.01
            return float(standard_price)
        return 0.01

    def get_selling_price(self, value):
        selling_price = value.get("Selling Price")
        if isinstance(selling_price, str):
            selling_price = selling_price.strip()
            selling_price = float(selling_price) if len(selling_price) > 0 else 0.0
        elif isinstance(selling_price, float):
            return selling_price
        elif isinstance(selling_price, int):
            return float(selling_price)
        return 0.0

    def get_or_create_book_language(self, value):
        book_language = value.get("Book Language")
        if book_language:
            book_language = book_language.strip()
            if len(book_language) > 0:
                book_language_obj = self.env["res.lang"].search(
                    [
                        ("name", "=ilike", book_language),
                        "|",
                        ("active", "=", False),
                        ("active", "=", True),
                    ], limit=1
                )
                if book_language_obj:
                    if not book_language_obj.active:
                        book_language_obj.active = True
                else:
                    book_language_obj = self.env["res.lang"].create(
                        {"name": book_language}
                    )
                return book_language_obj.id

    def get_or_create_product_format(self, value):
        product_format = value.get("Product Format")
        if product_format == "Cards":
            return "cards"
        elif product_format == "Audio":
            return "audio"
        elif product_format == "Book":
            return "book"
        elif product_format == "Bag":
            return "bag"
        elif product_format == "Hardware":
            return "hw"
        else:
            raise UserError(_("Product format is not valid, valid values for status are (Cards,  Audio, Book, Bag and Hardware)"))

    def get_or_create_status(self, value):
        status = value.get("Status")
        if status == "Upcoming":
            return "upcoming"
        elif status == "Available" or not status:
            return "available"
        else:
            raise UserError(_("Status is not valid, valid values for status are (Upcoming and Available)."))

    def get_publication_date(self, value):
        publication_date = value.get("Publication Date", False)
        if not publication_date:
            return False
        return publication_date

    def set_product_invoice_policy(self, product):
        if product.type in ["service", "consu"]:
            product.update({"invoice_policy": "order"})
        if product.type == "product":
            product.update({"invoice_policy": "delivery"})

    def validate_and_get_detailed_type(self, value):
        detailed_type = value.get("Product Type")
        if detailed_type:
            detailed_type = str(detailed_type).strip()
            # Map common variations to Odoo selection values
            # Odoo 19 type selection: [('consu', "Goods"), ('service', "Service"), ('combo', "Combo")]
            type_mapping = {
                "Service": "service",
                "service": "service",
                "Consumable": "consu",
                "consumable": "consu",
                "Storable Product": "consu",
                "storable product": "consu",
                "Goods": "consu",
                "goods": "consu",
                "product": "consu"
            }
            
            # Check case-insensitive mach first if exact match fails
            if detailed_type in type_mapping:
                 return type_mapping[detailed_type]
            
            # Try title case
            if detailed_type.title() in type_mapping:
                return type_mapping[detailed_type.title()]

            raise UserError(_("Product Type '%s' is not valid. Valid values are: Service, Consumable, Storable Product, Goods.") % detailed_type)
        else:
            raise UserError(_("Product Type is required."))

    def get_customer_taxes_ids(self, value):
        return self.env.company.account_sale_tax_id


    def get_vendor_taxes_ids(self, value):
        return self.env.company.account_purchase_tax_id


    def create_product(self, value):
        import_product = self.env.context.get("import_product")
        update_product = self.env.context.get("update_product")
        item_code = self.validate_and_get_item_code(value)
        
        # Validation step: check if duplicate exists in database (Barcode or Internal Reference)
        # Validation step: check if duplicate exists in database (Barcode or Internal Reference)
        self.env.cr.execute("SELECT id FROM product_product WHERE barcode = %s OR default_code = %s LIMIT 1", (str(item_code), str(item_code)))
        
        if self.env.cr.fetchone():
            return
        
        product = False

        product_vals = {
            "name": self.validate_and_get_item_name(value),
            "categ_id": self.get_item_category(value),
            "list_price": self.get_selling_price(value),
            "author_id": self.get_or_create_author(value),
            "publisher_id": self.get_or_create_publisher(value),
            "master_publisher_id": self.get_or_create_master_publisher(value),
            "uk_wholesaler_id": self.get_or_create_wholesaler(value, True),
            "non_uk_wholesaler_id": self.get_or_create_wholesaler(value, False),
            "publication_country_id": self.get_publication_country(value),
            "main_title": value.get("Main Title", False),
            "subtitle": value.get("Subtitle", False),
            "subject": value.get("Subject", False),
            "page": value.get("Pages", 0),
            "edition": value.get("Edition", False),
            "series": value.get("Series", False),
            "book_language_id": self.get_or_create_book_language(value),
            "product_format": self.get_or_create_product_format(value),
            "interest_age": value.get("interest_age", 0.0),
            "audience_readership": value.get("audience_readership"),
            "publication_date": self.get_publication_date(value),
            "status": self.get_or_create_status(value),
            "product_height": value.get("Height Along Spine In Mm", 0),
            "product_width": value.get("Width From Spine To Edge In Mm", 0),
            "weight": self.validate_and_get_weight(value),
            "type": self.validate_and_get_detailed_type(value),
            "sale_ok": value.get("Can be sold") or True,
            "purchase_ok": value.get("Can be purchased") or True,
            "available_in_pos": value.get("Available in POS") or True,
            "taxes_id": self.get_customer_taxes_ids(value),
            "supplier_taxes_id": self.get_vendor_taxes_ids(value)
        }
        if not product:
            if import_product and not update_product:
                product_vals.update({
                    "barcode": item_code,
                    "default_code": item_code, # Set internal reference as well
                    "standard_price": self.get_cost_price(value),
                    "purchase_method": 'purchase',
                })
                product = self.env["product.product"].create(product_vals)
                self.set_product_invoice_policy(product)
        else:
            if not import_product and update_product:
                product_vals.update({
                    "standard_price": self.get_cost_price(value)
                })
                product.write(product_vals)
                self.set_product_invoice_policy(product)

    def import_product(self):
        if not self.file:
            raise UserError(_("Please upload file!."))
        ext = os.path.splitext(self.file_name)
        ext = ext and len(ext) > 1 and ext[1] or ''
        if ext.lower() not in ['.xlsx', '.xls']:
            raise UserError(_('Selected file must in .xlsx format!'))
        fp = tempfile.NamedTemporaryFile(suffix=ext)
        fp.write(binascii.a2b_base64(self.file))
        fp.flush()
        fp.seek(0)
        
        # Try openpyxl first (handles xlsx)
        try:
            workbook = openpyxl.load_workbook(fp, data_only=True)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                row_vals = dict(zip(self.field_list_to_import, list(row)))
                if not row_vals.get("Item Code"):
                    continue
                self.create_product(row_vals)
        except UserError:
            raise
        except Exception as e_xlsx:
            # Fallback to xlrd (handles xls)
            try:
                workbook = xlrd.open_workbook(fp.name)
                sheet = workbook.sheet_by_index(0)
                for row_no in range(sheet.nrows):
                    if row_no != 0:
                        row_values = sheet.row_values(row_no)
                        # Re-implementing original xlrd cell logic:
                        row_cells = sheet.row(row_no)
                        val_counter = 0
                        row_val_list = []
                        for i in row_cells:
                            val = i.value
                            if i.ctype == xlrd.XL_CELL_DATE:
                                if val:
                                    val = datetime.datetime(
                                        *xlrd.xldate_as_tuple(val, workbook.datemode)
                                    )
                            row_val_list.append(val)
                            val_counter += 1
                        row_dict = dict(zip(self.field_list_to_import, row_val_list))
                        if not row_dict.get("Item Code"):
                            continue
                        self.create_product(row_dict)
            except UserError:
                raise
            except Exception as e_xls:
                raise UserError(_('Could not open file.\nOpenPyXL: %s\nXLRD: %s') % (str(e_xlsx), str(e_xls)))

    def download_all_products(self):
        self.env.cr.execute("""
                    SELECT id FROM product_template
                    WHERE active = TRUE
                    ORDER BY id DESC
                """)
        product_ids = [row[0] for row in self.env.cr.fetchall()]
        products_data = self.env['product.template'].sudo().browse(product_ids).read(
            ['barcode', 'name', 'display_name', 'standard_price']
        )
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            file_path = tmp_file.name
        workbook = xlsxwriter.Workbook(file_path)
        worksheet = workbook.add_worksheet('Report')
        title_format = workbook.add_format({'align': 'center', 'bold': True, 'font_size': 15})
        header_format = workbook.add_format({'bold': True, 'bg_color': '#D7E4BC'})
        date_format = workbook.add_format({'num_format': 'yyyy-mm-dd'})
        worksheet.merge_range('A2:D3', 'PRODUCTS LIST', title_format)
        worksheet.set_column('A:D', 30)
        worksheet.write('D4', f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d')}", date_format)
        headers = ['Barcode (ISBN)', 'Name', 'Display Name', 'Cost']
        worksheet.write_row('A6', headers, header_format)
        for row_idx, product in enumerate(products_data, start=6):  # Start from row 7 (0-indexed)
            worksheet.write(row_idx, 0, product.get('barcode') or '')
            worksheet.write(row_idx, 1, product.get('name') or '')
            worksheet.write(row_idx, 2, product.get('display_name') or '')
            worksheet.write(row_idx, 3, product.get('standard_price') or 0.0)
        workbook.close()
        with open(file_path, 'rb') as f:
            file_content = f.read()
        self.write({
            'sample_file': base64.b64encode(file_content),
            'file_name': 'product_report.xlsx',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/?model=import.product&id={self.id}&field=sample_file&download=true&filename={self.file_name}',
            'target': 'self',
        }
