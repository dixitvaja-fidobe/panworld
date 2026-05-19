import base64
import io
import csv
import openpyxl
import xlrd
from odoo import api, fields, models, _
from odoo.exceptions import UserError

class VendorPricelistImportWizard(models.TransientModel):
    _name = 'vendor.pricelist.import.wizard'
    _description = 'Import Vendor Pricelist Wizard'

    partner_id = fields.Many2one(
        'res.partner',
        string='Vendor',
        required=True,
    )
    currency_id = fields.Many2one(
        'res.currency',
        string='Currency',
        required=True,
        default=lambda self: self.env.company.currency_id,
    )
    csv_file = fields.Binary(string='File', required=True)
    filename = fields.Char(string='Filename')

    def action_download_template(self):
        output = io.BytesIO()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Pricelist Template"
        
        # Headers
        headers = ['ISBN', 'Price', 'Discount']
        sheet.append(headers)
        
        # Sample Data
        sheet.append(['9780123456789', 50.0, 10.0])
        sheet.append(['1234567890123', 75.5, 0.0])
        
        workbook.save(output)
        file_content = output.getvalue()
        
        attachment = self.env['ir.attachment'].create({
            'name': 'vendor_pricelist_template.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(file_content),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'new',
        }

    def action_import(self):
        self.ensure_one()
        if not self.csv_file:
            raise UserError(_('Please upload a file.'))

        file_data = base64.b64decode(self.csv_file)
        rows = []

        import zipfile
        filename = (self.filename or '').lower()
        is_zip = file_data.startswith(b'PK\x03\x04')
        is_xls = file_data.startswith(b'\xd0\xcf\x11\xe0')

        # 1. Attempt to read as Excel (XLSX)
        if is_zip or filename.endswith('.xlsx'):
            try:
                workbook = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
                sheet = workbook.active
                rows = list(sheet.iter_rows(values_only=True))
            except Exception as e:
                # Check if it's actually an ODS file
                if is_zip:
                    try:
                        with zipfile.ZipFile(io.BytesIO(file_data)) as z:
                            if 'mimetype' in z.namelist() and b'opendocument.spreadsheet' in z.read('mimetype'):
                                raise UserError(_("ODS format is not supported. Please save your file as Excel (.xlsx) or CSV."))
                    except UserError: raise
                    except Exception: pass
                
                # If it was supposed to be a ZIP/XLSX, don't fall back to CSV
                if filename.endswith('.xlsx') or is_zip:
                    raise UserError(_("Could not read the Excel file. Please ensure it is a valid .xlsx file. Error: %s") % str(e))

        # 2. Attempt to read as Excel (XLS)
        if not rows and (is_xls or filename.endswith('.xls')):
            try:
                workbook = xlrd.open_workbook(file_contents=file_data)
                sheet = workbook.sheet_by_index(0)
                rows = []
                for row_idx in range(sheet.nrows):
                    rows.append(sheet.row_values(row_idx))
            except Exception as e:
                if filename.endswith('.xls') or is_xls:
                    raise UserError(_("Could not read the Excel file. Please ensure it is a valid .xls file. Error: %s") % str(e))

        # 3. Fallback to CSV (Only if not already read and not a binary Excel file)
        if not rows:
            if is_zip or is_xls:
                raise UserError(_("The file is a binary archive (Excel/ODS) but could not be read. CSV fallback is not possible for binary data."))
            
            try:
                # Detect encoding and handle BOM
                try:
                    decoded_data = file_data.decode('utf-8-sig')
                except UnicodeDecodeError:
                    decoded_data = file_data.decode('latin-1')
                
                # Clean up data
                decoded_data = decoded_data.replace('\x00', '')
                
                # Basic delimiter detection
                delimiter = ','
                first_line = decoded_data.split('\n')[0] if decoded_data else ''
                if ';' in first_line and ',' not in first_line:
                    delimiter = ';'
                elif '\t' in first_line:
                    delimiter = '\t'
                
                f = io.StringIO(decoded_data, newline='')
                
                # Use Sniffer for more robust detection if possible
                try:
                    dialect = csv.Sniffer().sniff(decoded_data[:2048])
                    f.seek(0)
                    reader = csv.reader(f, dialect)
                except Exception:
                    f.seek(0)
                    reader = csv.reader(f, delimiter=delimiter)
                    
                rows = list(reader)
            except Exception as e:
                raise UserError(_('The file could not be read as Excel or CSV. Error: %s') % str(e))

        if not rows:
            raise UserError(_('The file is empty or could not be read.'))
            
        # Normalize rows: ensure all elements are strings or primitives, handle None
        # Headers should be determined from the first non-empty row?
        # Let's assume first row is header
        
        header = [str(h).lower().strip() if h else '' for h in rows[0]]
        
        # Identify columns
        try:
            isbn_idx = -1
            price_idx = -1
            discount_idx = -1
            
            for idx, h in enumerate(header):
                if h in ['isbn', 'barcode', 'code']:
                    isbn_idx = idx
                elif h in ['price', 'cost', 'unit price']:
                    price_idx = idx
                elif h in ['discount', 'disc', 'disc %', 'discount %']:
                    discount_idx = idx
            
            if isbn_idx == -1 or price_idx == -1:
                col_names = [h for h in header if h]
                raise UserError(_('Column headers "ISBN", "Price" and "Discount" are required.\nFound columns: %s') % ', '.join(col_names))
                
        except ValueError:
            raise UserError(_('Could not identify required columns.'))

        rows_data = []
        isbns_to_check = set()
        
        for row_idx, row in enumerate(rows[1:], start=2):
            if not row: continue
            
            # Safely access list index
            row_len = len(row)
            if row_len <= isbn_idx or row_len <= price_idx:
                continue # Skip malformed rows
                
            isbn_val = row[isbn_idx]
            if isinstance(isbn_val, float) and isbn_val % 1 == 0:
                isbn = str(int(isbn_val)).strip()
            else:
                isbn = str(isbn_val).strip() if isbn_val else ''
                
            price_val = row[price_idx]
            
            price = 0.0
            if price_val:
                try:
                    # Clean currency symbols if CSV string
                    if isinstance(price_val, str):
                        price_val = price_val.replace('$', '').replace('€', '').replace(',', '')
                    price = float(price_val)
                except ValueError:
                    pass 
            
            discount = 0.0
            if discount_idx != -1 and len(row) > discount_idx:
                disc_val = row[discount_idx]
                if disc_val:
                    try:
                        if isinstance(disc_val, str):
                             disc_val = disc_val.replace('%', '').replace(',', '')
                        discount = float(disc_val)
                    except ValueError:
                        pass 
            
            if isbn:
                rows_data.append({
                    'isbn': isbn,
                    'price': price,
                    'discount': discount,
                    'row_idx': row_idx
                })
                isbns_to_check.add(isbn)

        if not rows_data:
             raise UserError(_('No valid data found to import.'))

        # Step 1: Validation - Check for missing ISBNs
        products = self.env['product.product'].search([
            '|', ('default_code', 'in', list(isbns_to_check)),
                 ('barcode', 'in', list(isbns_to_check))
        ])
        
        found_map = {} # isbn -> product_id
        for p in products:
            if p.default_code:
                found_map[p.default_code] = p.id
            if p.barcode:
                found_map[p.barcode] = p.id

        missing_isbns = [isbn for isbn in isbns_to_check if isbn not in found_map]
        
        if missing_isbns:
            shown_missing = missing_isbns[:20]
            msg = _('The following ISBNs/Barcodes were not found in the system:\n%s') % '\n'.join(shown_missing)
            if len(missing_isbns) > 20:
                msg += _('\n... and %s more.') % (len(missing_isbns) - 20)
            raise UserError(msg)

        # Step 2: Import & Deduplication
        SupplierInfo = self.env['product.supplierinfo']
        created_count = 0
        
        for data in rows_data:
            product_id = found_map.get(data['isbn'])
            product = self.env['product.product'].browse(product_id)
            
            # Check for existing pricelist
            existing = SupplierInfo.search([
                ('partner_id', '=', self.partner_id.id),
                ('product_tmpl_id', '=', product.product_tmpl_id.id),
                ('currency_id', '=', self.currency_id.id),
                ('price', '=', data['price']),
                ('min_qty', '=', 0.0),
            ], limit=1)
            
            if existing:
                continue
            
            # Create new pricelist
            SupplierInfo.create({
                'partner_id': self.partner_id.id,
                'product_tmpl_id': product.product_tmpl_id.id,
                'product_id': product.id,
                'currency_id': self.currency_id.id,
                'price': data['price'],
                'discount': data['discount'],
                'min_qty': 0.0,
            })
            created_count += 1

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import Successful'),
                'message': _('%d vendor pricelists created.') % created_count,
                'type': 'success',
                'sticky': False,
                'next': {'type': 'ir.actions.act_window_close'},
            }
        }
