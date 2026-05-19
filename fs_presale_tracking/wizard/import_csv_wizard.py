import base64
import csv
import io
import openpyxl
import xlrd
import xlsxwriter
from odoo import api, fields, models, _
from odoo.exceptions import UserError


class PresaleTrackingImportWizard(models.TransientModel):
    _name = 'presale.tracking.import.wizard'
    _description = 'Import CSV Wizard'

    tracking_id = fields.Many2one(
        'presale.tracking',
        string='Presale Tracking',
        required=True,
    )
    csv_file = fields.Binary(string='File', required=True)
    csv_filename = fields.Char(string='Filename')
    delimiter = fields.Selection([
        (',', 'Comma (,)'),
        (';', 'Semicolon (;)'),
        ('\t', 'Tab'),
    ], string='Delimiter', default=',', required=True, help="Only for CSV files")
    is_csv = fields.Boolean(compute='_compute_is_csv')
    csv_file = fields.Binary(string='Upload your file', required=False) # Changed to required=False to allow export-only

    @api.depends('csv_filename')
    def _compute_is_csv(self):
        for record in self:
            record.is_csv = record.csv_filename and record.csv_filename.lower().endswith('.csv')

    clear_existing = fields.Boolean(
        string='Clear Existing Lines',
        default=False,
        help='If checked, all existing lines will be removed before import',
    )

    def action_import_lines(self):
        self.ensure_one()

        if not self.csv_file:
            raise UserError(_('Please upload a file.'))

        file_content = base64.b64decode(self.csv_file)
        extension = (self.csv_filename or '').split('.')[-1].lower()

        rows_data = []
        
        if extension == 'csv':
            # Handle CSV
            try:
                csv_data = file_content.decode('utf-8')
            except UnicodeDecodeError:
                csv_data = file_content.decode('latin-1')
            
            reader = csv.DictReader(io.StringIO(csv_data), delimiter=self.delimiter)
            for row in reader:
                rows_data.append(row)
        
        elif extension == 'xlsx':
            # Handle XLSX
            try:
                wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
                sheet = wb.active
                headers = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
                for row_idx in range(2, sheet.max_row + 1):
                    row_values = [cell.value for cell in sheet[row_idx]]
                    row_dict = dict(zip(headers, row_values))
                    if any(row_dict.values()): # Skip empty rows
                        rows_data.append(row_dict)
            except Exception as e:
                raise UserError(_("Error reading XLSX file: %s") % str(e))
        
        elif extension == 'xls':
            # Handle XLS
            try:
                wb = xlrd.open_workbook(file_contents=file_content)
                sheet = wb.sheet_by_index(0)
                headers = [str(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols)]
                for row_idx in range(1, sheet.nrows):
                    row_values = [sheet.cell_value(row_idx, col) for col in range(sheet.ncols)]
                    row_dict = dict(zip(headers, row_values))
                    if any(row_dict.values()): # Skip empty rows
                        rows_data.append(row_dict)
            except Exception as e:
                raise UserError(_("Error reading XLS file: %s") % str(e))
        
        else:
            raise UserError(_("Unsupported file extension: %s. Please upload a .csv, .xlsx or .xls file.") % extension)

        # Clear existing lines if requested
        if self.clear_existing:
            self.env.cr.execute("DELETE FROM presale_tracking_line WHERE tracking_id=%s", (self.tracking_id.id,))
            self.tracking_id.invalidate_recordset(['line_ids'])

        lines_to_create = []
        updates_performed = 0
        row_count = 0
        
        for row in rows_data:
            row_count += 1
            
            line_id = row.get('ID') or row.get('id') or False
            isbn_val = row.get('ISBN') or row.get('isbn') or row.get('Isbn') or ''
            
            if isinstance(isbn_val, float) and isbn_val.is_integer():
                isbn = str(int(isbn_val))
            else:
                isbn = str(isbn_val).strip()

            title = str(row.get('TITLE') or row.get('title') or row.get('Title') or '').strip()

            if not isbn and not line_id:
                continue

            # Get QTY
            qty_header = next((h for h in ['QTY', 'qty', 'Qty', 'Quantity'] if h in row), None)
            qty = 0
            if qty_header:
                qty_val = row[qty_header]
                if qty_val is not None and not (isinstance(qty_val, str) and not qty_val.strip()):
                    try:
                        qty_float = float(qty_val)
                        if not qty_float.is_integer():
                             raise UserError(_("Line %s: Quantity '%s' is not an integer") % (row_count, qty_val))
                        qty = int(qty_float)
                    except (ValueError, TypeError):
                        raise UserError(_("Line %s: Quantity '%s' is not an integer") % (row_count, qty_val))

            # Get List Price and Discount
            price_header = next((h for h in ['List Price', 'list_price', 'Price'] if h in row), None)
            list_price = 0.0
            if price_header:
                try:
                    list_price = float(row[price_header] or 0.0)
                except (ValueError, TypeError):
                    pass
            
            disc_header = next((h for h in ['Purchase Discount', 'pub_disc', 'Discount'] if h in row), None)
            discount = 0.0
            if disc_header:
                try:
                    discount = float(row[disc_header] or 0.0)
                except (ValueError, TypeError):
                    pass
            
            sales_disc_header = next((h for h in ['Sales Discount', 'disc_percent'] if h in row), None)
            sales_discount = 0.0
            if sales_disc_header:
                try:
                    sales_discount = float(row[sales_disc_header] or 0.0)
                except (ValueError, TypeError):
                    pass

            vals = {
                'tracking_id': self.tracking_id.id,
                'isbn': isbn,
                'title': title,
                'qty': qty,
                'list_price': list_price,
                'pub_disc': discount,
                'disc_percent': sales_discount,
            }

            if line_id:
                try:
                    line_id_int = int(float(line_id))
                    existing_line = self.env['presale.tracking.line'].browse(line_id_int)
                    if existing_line.exists() and existing_line.tracking_id == self.tracking_id:
                        # Update existing line
                        # We only update if ISBN matches or if it's explicitly identifying this line
                        update_vals = {}
                        if isbn: update_vals['isbn'] = isbn
                        if title: update_vals['title'] = title
                        if qty_header: update_vals['qty'] = qty
                        if price_header: update_vals['list_price'] = list_price
                        if disc_header: update_vals['pub_disc'] = discount
                        if sales_disc_header: update_vals['disc_percent'] = sales_discount
                        
                        existing_line.write(update_vals)
                        updates_performed += 1
                        continue
                except (ValueError, TypeError):
                    pass
            
            # If no ID or ID not found, treat as new line
            vals['sequence'] = row_count * 10
            lines_to_create.append(vals)

        # Batch lookup products for new lines
        if lines_to_create:
            all_isbns = [r['isbn'] for r in lines_to_create if r.get('isbn')]
            products = self.env['product.product'].search([('default_code', 'in', all_isbns)])
            product_map = {p.default_code: p.id for p in products}

            for line_vals in lines_to_create:
                if line_vals.get('isbn') in product_map:
                    line_vals['product_id'] = product_map[line_vals['isbn']]
            
            created_lines = self.env['presale.tracking.line'].create(lines_to_create)
            
            # Explicitly re-write the provided price and discount values to override 
            # any computed defaults inherited from the product's vendor/seller_id
            for line, vals in zip(created_lines, lines_to_create):
                price_update = {}
                if 'list_price' in vals:
                    price_update['list_price'] = vals['list_price']
                if 'pub_disc' in vals:
                    price_update['pub_disc'] = vals['pub_disc']
                if 'disc_percent' in vals:
                    price_update['disc_percent'] = vals['disc_percent']
                
                if price_update:
                    line.write(price_update)

        # Save file as attachment
        mimetypes = {
            'csv': 'text/csv',
            'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'xls': 'application/vnd.ms-excel',
        }
        self.env['ir.attachment'].create({
            'name': self.csv_filename or 'imported_data',
            'type': 'binary',
            'datas': self.csv_file,
            'res_model': 'presale.tracking',
            'res_id': self.tracking_id.id,
            'mimetype': mimetypes.get(extension, 'application/octet-stream'),
        })

        if self.tracking_id.state == 'draft' and self.tracking_id.request_file:
            self.tracking_id.action_in_progress()

        message = _('%d lines created. %d lines updated.') % (len(lines_to_create), updates_performed)
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import Successful'),
                'message': message,
                'type': 'success',
                'sticky': False,
                'next': {'type': 'ir.actions.act_window_close'},
            }
        }

    def action_export_existing(self):
        self.ensure_one()
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        sheet = workbook.add_worksheet('Tracking Lines')
        
        # Formats
        header_format = workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#D3D3D3'})
        locked_format = workbook.add_format({'locked': True, 'bg_color': '#EEEEEE'})
        unlocked_format = workbook.add_format({'locked': False})

        # Columns
        headers = ['ID', 'ISBN', 'TITLE', 'QTY', 'List Price', 'Purchase Discount', 'Sales Discount']
        for col, header in enumerate(headers):
            sheet.write(0, col, header, header_format)
            sheet.set_column(col, col, 15)

        # Write lines
        for row_idx, line in enumerate(self.tracking_id.line_ids, 1):
            sheet.write(row_idx, 0, line.id, locked_format)
            sheet.write(row_idx, 1, line.isbn or '', unlocked_format)
            sheet.write(row_idx, 2, line.title or '', unlocked_format)
            sheet.write(row_idx, 3, line.qty, unlocked_format)
            sheet.write(row_idx, 4, line.list_price, unlocked_format)
            sheet.write(row_idx, 5, line.pub_disc, unlocked_format)
            sheet.write(row_idx, 6, line.disc_percent, unlocked_format)

        # Protect sheet (but allow selection)
        sheet.protect()
        
        workbook.close()
        output.seek(0)
        
        file_content = output.read()
        
        attachment = self.env['ir.attachment'].create({
            'name': 'Existing_Tracking_Lines_%s.xlsx' % (self.tracking_id.name or 'Export'),
            'type': 'binary',
            'datas': base64.b64encode(file_content),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }
