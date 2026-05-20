# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) Fidobe Solutions LLC.
#
#    For Module Support : dev3@fidobe.com
#
#############################################################################
import binascii
import logging
import os
import tempfile

import xlsxwriter
from datetime import datetime
from odoo.tools import DEFAULT_SERVER_DATE_FORMAT as DATE_FORMAT

from odoo import _, fields, models, exceptions
from odoo.exceptions import UserError,ValidationError
from collections import defaultdict

from odoo import _, fields, models,api

_logger = logging.getLogger(__name__)


try:
    import xlrd
except ImportError:
    _logger.debug("Cannot `import xlrd`.")

try:
    import openpyxl
except ImportError:
    _logger.debug("Cannot `import openpyxl`.")
try:
    import base64
except ImportError:
    _logger.debug("Cannot `import base64`.")


class ImportDataWizard(models.TransientModel):
    _name = "import.data.wizard"
    _description = "Import Data Wizard"

    data_import_file = fields.Binary(string="Select File")
    sample_file = fields.Binary("Download File")
    file_name = fields.Char(string="File Name")
    model_name = fields.Char(string="Model Name")
    model_id = fields.Integer(string="Model Id")
    show_export_button= fields.Boolean(string='Show Export Button', compute='_compute_show_export_button', store=True)
    show_po_export_button = fields.Boolean(string='Show PO Export Button', compute='_compute_show_po_export_button', store=True)
    show_rma_export_button = fields.Boolean(string='Show RMA Export Button', compute='_compute_show_rma_export_button', store=True)
    without_isbn = fields.Boolean(default=False, string="Import Without ISBN")

    @api.depends('model_name')
    def _compute_show_export_button(self):
        for rec in self:
            if rec.env.context.get('active_model') in ['purchase.order', 'account.move', 'stock.picking', 'product.pricelist']:
                rec.show_export_button = True
            else:
                rec.show_export_button = False

    @api.depends('model_name')
    def _compute_show_po_export_button(self):
        for rec in self:
            if rec.env.context.get('active_model') == 'purchase.order' or rec.model_name == 'purchase.order':
                rec.show_po_export_button = True
            else:
                rec.show_po_export_button = False

    @api.depends('model_name')
    def _compute_show_rma_export_button(self):
        for rec in self:
            if rec.env.context.get('active_model') == 'rma.ret.mer.auth' or rec.model_name == 'rma.ret.mer.auth':
                rec.show_rma_export_button = True
            else:
                rec.show_rma_export_button = False
    def download_sample_file(self):
        """Method to download sample file for picking line import"""
        name_of_file = "sample_with_import_data.xls"
        fp = tempfile.NamedTemporaryFile(delete=False, suffix=".xls")
        path_list = list(os.path.split(fp.name))
        path_list[len(path_list) - 1] = name_of_file
        file_path = os.path.join(*path_list)
        workbook = xlsxwriter.Workbook(file_path)
        worksheet = workbook.add_worksheet()
        counter = 0

        context = self.env.context
        # extra: to fix issue of import
        model_name = self.model_name if self.model_name else context.get("active_model")
        model_id = self.model_id if self.model_id else context.get('active_id')
        model_rec = self.env[model_name].browse([model_id])

        field_list_to_import = []
        if model_name == "stock.picking":
            field_list_to_import = ["id", "isbn", "bundle_isbn", "demand", "done", "cancel_reason"]
            row = 1
            for line in model_rec.move_ids:
                cancel_reason_selection = [l[0] for l in line._fields['cancel_reason'].selection]
                worksheet.data_validation(row, 3, row, 3, {'validate': 'list', 'source': cancel_reason_selection})
                row += 1
        if model_name == "account.move" or  self.env.context.get('active_model') == 'account.move':
            field_list_to_import = ["id", "isbn", "quantity", "list_price", "discount", "tax_ids", "analytic_distribution"]
        if model_name == "purchase.order":
            field_list_to_import = ["Id", "Isbn", "RFQ Quantity", 'To Be Cancelled', "To Be Received QTY",
                                    "Po List Price", "Po Discount",
                                    "Po Target Price", "Bill List Price", "Bill Discount", "Sales Ref- SOR Odoo", "Remarks",
                                    'Cancel Reason']
        if model_name == "purchase.request":
            field_list_to_import = ["isbn", "qty", "estimated cost"]
        if model_name == "product.pricelist":
            field_list_to_import = ["id", "isbn", "applied_on", "compute_price", "fixed_price", "percent_price", "min_quantity", "date_start", "date_end"]
        if not context.get("sale_direct") and model_name == "rma.ret.mer.auth":
            field_list_to_import = ["isbn", "source_location_id", "source_parent_id", "destination_location_id",
                                    "destination_parent_id", "order_quantity", "delivered_quantity", "refund_qty",
                                    "price_unit", "tax_id", "refund_price"]
        if context.get("sale_direct") and model_name == "rma.ret.mer.auth":
            field_list_to_import = ["isbn", "source_location_id", "source_parent_id", "destination_location_id",
                                    "destination_parent_id", "order_quantity", "delivered_quantity", "refund_qty",
                                    "price_unit", "tax_id", "refund_price",
                                    "lco", "mco", "sco", "oco", "tco"]

        header_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'bold': True, 'border': 1})

        for i in field_list_to_import:
            worksheet.write(0, counter, i, header_format)
            counter += 1
        workbook.close()
        export_id = base64.b64encode(open(file_path, "rb+").read())
        self.write({"sample_file": export_id, "file_name": name_of_file})
        return {
            "name": "Update Picking Lines",
            "view_mode": "form",
            "res_id": self.id,
            "res_model": "import.data.wizard",
            "view_type": "form",
            "type": "ir.actions.act_window",
            "target": "new",
        }

    def validate_and_get_tax(self, value):
        """Method to get Tax"""
        tax_ids = value.get("tax_ids")
        if not tax_ids:
            return False
        tax_ids = str(tax_ids).split(",")
        return tax_ids

    def import_data(self):
        """Method to import delivery order lines"""
        if not self.data_import_file:
            raise UserError(_("Please upload file!."))
        context = self.env.context

        # extra: to fix issue of import
        model_name = self.model_name if self.model_name else context.get("active_model")
        model_id = self.model_id if self.model_id else context.get('active_id')
        model_rec = self.env[model_name].browse([model_id])
        fp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        fp.write(binascii.a2b_base64(self.data_import_file))
        fp.seek(0)
        invoice_line_commands = []
        workbook = None
        try:
             # Try opening as xlsx with openpyxl (read_only for large files)
            workbook = openpyxl.load_workbook(fp.name, read_only=True, data_only=True)
            sheet = workbook.active
            headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            is_xlsx = True
        except Exception:
            # Fallback to xlrd for xls
            is_xlsx = False
            sheet = xlrd.open_workbook(fp.name).sheet_by_index(0)
            headers = [sheet.cell_value(0, i) for i in range(sheet.ncols)]

        file_headers = [str(x or '').lower().strip() for x in headers]
        file_headers = [h for h in file_headers if h]

        # Dynamically find ISBN column index
        isbn_index = -1
        for i, h in enumerate(file_headers):
            if 'isbn' in h:
                isbn_index = i
                break

        codes = []
        if isbn_index != -1:
            if is_xlsx:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if len(row) > isbn_index and row[isbn_index]:
                        codes.append(str(row[isbn_index]).split('.')[0])
            else: # xlrd
                for rownum in range(1, sheet.nrows): # Skip header
                    if sheet.ncols > isbn_index:
                        codes.append(str(sheet.cell_value(rownum, isbn_index)).split('.')[0])
        if not is_xlsx:
            # We already populated codes in xlrd block above
             pass
        # For xlsx, codes are already populated in try block

        # --- ULTRA-FAST PRE-CHECK & CACHE ---
        msg = ""
        product_cache = self._build_product_cache(codes) if codes else {}
        if codes:
            found_codes = set(product_cache.keys())
            missing = [c for c in codes if c not in found_codes]
            if missing:
                msg = ", ".join([f"[{c}]" for c in missing[:50]])
                if len(missing) > 50: msg += "..."
        
        if msg and not self.without_isbn:
            raise UserError(_("Products not found:\n %s", msg))
        list_records = []
        field_list_to_import = ["id","isbn","bundle_isbn","demand","done","cancel_reason"]
        if model_name == "account.move":
            expected_headers = ["id", "isbn", "description","quantity", "list_price", "discount", "tax_ids",
                                    "account_id", "analytic_distribution"]
            expected_headers_no_id = ["isbn", "description","quantity", "list_price", "discount", "tax_ids",
                                    "account_id", "analytic_distribution"]
            if model_rec.move_type == "entry":
                expected_headers = ["id", "account_id", "partner_id", "label", "analytic_distribution", "debit", "credit"]
                expected_headers_no_id = ["account_id", "partner_id", "label", "analytic_distribution", "debit", "credit"]
            
            if file_headers == expected_headers:
                field_list_to_import = expected_headers
            elif file_headers == expected_headers_no_id:
                field_list_to_import = expected_headers_no_id
            else:
                raise UserError(
                    _("File headers must match the sample file's headers. \nExpected: %s \nFound: %s") % (expected_headers, file_headers))
            if is_xlsx:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    invoice_value = dict(zip(field_list_to_import, row))
                    # For Journal Entries, we don't have isbn, so we check for account_id or id
                    if model_rec.move_type == "entry":
                        if not invoice_value.get('id') and not invoice_value.get('account_id'):
                            continue
                    elif not invoice_value.get('id') and not invoice_value.get('isbn'):
                        continue
                    
                    id = int(invoice_value.get('id')) if invoice_value.get('id') else False
                    isbn = str(invoice_value.get('isbn', '')).split('.')[0]
                    value = {}
                    product = self.env['product.product'].search([('barcode', '=', isbn)], limit=1)
                    # Build base values
                    for field in ['quantity', 'discount']:
                        if invoice_value.get(field):
                            value[field] = invoice_value[field]
                    list_price = float(invoice_value.get('list_price')) if self.is_number(invoice_value.get('list_price')) else 0
                    discount = float(invoice_value.get('discount')) if self.is_number(invoice_value.get('discount')) else 0
                    value['discount'] = discount
                    value['price_unit'] = float(list_price)
                    value['list_price'] = list_price - (list_price * discount / 100.0) if discount else value['price_unit']
                    tax_names = self.validate_and_get_tax(invoice_value) if invoice_value.get('tax_ids') else []
                    if tax_names:
                        invoice_type = model_rec.move_type
                        if invoice_type in ['out_invoice', 'out_refund']:
                            tax_type = 'sale'
                        elif invoice_type in ['in_invoice', 'in_refund']:
                            tax_type = 'purchase'
                        else:
                            tax_type = False
                        domain = [("name", "in", tax_names)]
                        if tax_type:
                            domain.append(("type_tax_use", "=", tax_type))
                        tax_names = [tax.strip() for tax in tax_names]
                        taxes = self.env["account.tax"].search(domain)
                        found_tax_names = set(taxes.mapped('name'))
                        missing_taxes = set(tax_names) - found_tax_names
                        if missing_taxes:
                            raise UserError(_("Taxes %s not found in system!", ', '.join(missing_taxes)))
                        value['tax_ids'] = [(6, 0, taxes.ids)]
                    # Partner
                    if invoice_value.get('partner_id'):
                        partner_name = invoice_value['partner_id']
                        partner = self.env["res.partner"].search([("name", "=", partner_name)], limit=1)
                        if not partner:
                            raise UserError(_("Partner %s not found in system!", invoice_value.get('partner_id')))
                        value['partner_id'] = partner.id
                    # Account
                    if invoice_value.get('account_id'):
                        account_code = str(invoice_value['account_id']).strip().split()[0]
                        account = self.env["account.account"].search([
                            ("code", "=", account_code),
                            ("company_ids", "in", [model_rec.company_id.id])
                        ], limit=1)
                        if not account:
                            raise UserError(_("Account %s not found in system!", invoice_value.get('account_id')))
                        value['account_id'] = account.id
                    # Analytic Account
                    if invoice_value.get('analytic_distribution'):
                        analytic_name = invoice_value['analytic_distribution'].strip()
                        analytic_account = self.env["account.analytic.account"].search([("name", "=", analytic_name)],
                                                                                        limit=1)
                        if not analytic_account:
                            raise UserError(_("Analytic Account %s not found in system!", analytic_name))
                        analytic_data = value.get('analytic_distribution', {}) or {}
                        analytic_data[str(analytic_account.id)] = 100.0
                        value['analytic_distribution'] = analytic_data
                    description = product.display_name if product else ''
                    value['name'] = invoice_value['description'].strip() if invoice_value.get('description') else description
                    if invoice_value.get('label') and model_rec.move_type=="entry":
                        value['name'] = invoice_value['label'].strip()
                    value['debit'] = invoice_value['debit'] if invoice_value.get('debit') else 0
                    value['credit'] = invoice_value['credit'] if invoice_value.get('credit') else 0
                    # Check for existing line
                    existing_line = self.env['account.move.line'].browse(id)
                    if existing_line:
                        invoice_line_commands.append((1, existing_line.id, value))
                        # invoice.write({'invoice_line_ids': [(1, existing_line.id, value)]})
                    else:
                        value.update({'move_id': model_rec.id,
                                      'product_id': product.id if product else False,})
                        # invoice.write({'invoice_line_ids': [(0, 0, value)]})
                        invoice_line_commands.append((0, 0, value))
            
            else: # xlrd
                for row_no in range(1, sheet.nrows): # Skip header
                    invoice_value = dict(zip(field_list_to_import, sheet.row_values(row_no)))
                    # For Journal Entries, we don't have isbn, so we check for account_id or id
                    if model_rec.move_type == "entry":
                        if not invoice_value.get('id') and not invoice_value.get('account_id'):
                            continue
                    elif not invoice_value.get('id') and not invoice_value.get('isbn'):
                        continue

                    id = int(invoice_value.get('id')) if invoice_value.get('id') else False
                    isbn = str(invoice_value.get('isbn', '')).split('.')[0]
                    value = {}
                    product = self.env['product.product'].search([('barcode', '=', isbn)], limit=1)
                    # Build base values
                    for field in ['quantity', 'discount']:
                        if invoice_value.get(field):
                            value[field] = invoice_value[field]
                    list_price = float(invoice_value.get('list_price')) if self.is_number(invoice_value.get('list_price')) else 0
                    discount = float(invoice_value.get('discount')) if self.is_number(invoice_value.get('discount')) else 0
                    value['discount'] = discount
                    value['price_unit'] = float(list_price)
                    value['list_price'] = list_price - (list_price * discount / 100.0) if discount else value['price_unit']
                    tax_names = self.validate_and_get_tax(invoice_value) if invoice_value.get('tax_ids') else []
                    if tax_names:
                        invoice_type = model_rec.move_type
                        if invoice_type in ['out_invoice', 'out_refund']:
                            tax_type = 'sale'
                        elif invoice_type in ['in_invoice', 'in_refund']:
                            tax_type = 'purchase'
                        else:
                            tax_type = False
                        domain = [("name", "in", tax_names)]
                        if tax_type:
                            domain.append(("type_tax_use", "=", tax_type))
                        tax_names = [tax.strip() for tax in tax_names]
                        taxes = self.env["account.tax"].search(domain)
                        found_tax_names = set(taxes.mapped('name'))
                        missing_taxes = set(tax_names) - found_tax_names
                        if missing_taxes:
                            raise UserError(_("Taxes %s not found in system!", ', '.join(missing_taxes)))
                        value['tax_ids'] = [(6, 0, taxes.ids)]
                    # Partner
                    if invoice_value.get('partner_id'):
                        partner_name = invoice_value['partner_id']
                        partner = self.env["res.partner"].search([("name", "=", partner_name)], limit=1)
                        if not partner:
                            raise UserError(_("Partner %s not found in system!", invoice_value.get('partner_id')))
                        value['partner_id'] = partner.id
                    # Account
                    if invoice_value.get('account_id'):
                        account_code = invoice_value['account_id'].strip().split()[0]
                        account = self.env["account.account"].search([
                            ("code", "=", account_code),
                            ("company_ids", "in", [model_rec.company_id.id])
                        ], limit=1)
                        if not account:
                            raise UserError(_("Account %s not found in system!", invoice_value.get('account_id')))
                        value['account_id'] = account.id
                    # Analytic Account
                    if invoice_value.get('analytic_distribution'):
                        analytic_name = invoice_value['analytic_distribution'].strip()
                        analytic_account = self.env["account.analytic.account"].search([("name", "=", analytic_name)],
                                                                                       limit=1)
                        if not analytic_account:
                            raise UserError(_("Analytic Account %s not found in system!", analytic_name))
                        analytic_data = value.get('analytic_distribution', {}) or {}
                        analytic_data[str(analytic_account.id)] = 100.0
                        value['analytic_distribution'] = analytic_data
                    description = product.display_name if product else ''
                    value['name'] = invoice_value['description'].strip() if invoice_value.get('description') else description
                    if invoice_value.get('label') and model_rec.move_type=="entry":
                        value['name'] = invoice_value['label'].strip()
                    value['debit'] = invoice_value['debit'] if invoice_value.get('debit') else 0
                    value['credit'] = invoice_value['credit'] if invoice_value.get('credit') else 0
                    # Check for existing line
                    existing_line = self.env['account.move.line'].browse(id)
                    if existing_line:
                        invoice_line_commands.append((1, existing_line.id, value))
                        # invoice.write({'invoice_line_ids': [(1, existing_line.id, value)]})
                    else:
                        value.update({'move_id': model_rec.id,
                                      'product_id': product.id if product else False,})
                        # invoice.write({'invoice_line_ids': [(0, 0, value)]})
                        invoice_line_commands.append((0, 0, value))
            if invoice_line_commands:
                model_rec.write({'invoice_line_ids': invoice_line_commands})


        elif model_name == "purchase.order" and file_headers[-1] == 'bill discount':
            field_list_to_import = ['isbn', 'to be received qty', 'bill list price', 'bill discount']
            if file_headers != field_list_to_import:
                raise UserError(
                    _("File headers must match the sample file's headers. \nExpected: %s \nFound: %s") % (field_list_to_import, file_headers))
        elif model_name == "purchase.order":
            field_list_to_import = ["Id", "Isbn", "RFQ Quantity", 'To Be Cancelled', "To Be Received QTY",
                 "Po List Price", "Po Discount",
                 "Po Target Price", "Bill List Price", "Bill Discount", "Sales Ref- SOR Odoo", "Remarks",
                 'Cancel Reason']

            field_list_to_import = [x.lower() for x in field_list_to_import]
            if file_headers != field_list_to_import:
                raise UserError(
                    _("File headers must match the sample file's headers. \nExpected: %s \nFound: %s") % (field_list_to_import, file_headers))
            for line in model_rec.order_line:
                if line.product_id.barcode not in codes and model_rec.state in ("sent", "draft"):
                    line.unlink()
        elif model_name == "purchase.request":
            field_list_to_import = ["isbn", "qty", "estimated cost"]
            if file_headers != field_list_to_import:
                raise UserError(
                    _("File headers must match the sample file's headers. \nExpected: %s \nFound: %s") % (field_list_to_import, file_headers))
        elif model_name == "product.pricelist":
            field_list_to_import = ["id", "isbn", "applied_on", "compute_price", "fixed_price", "percent_price", "min_quantity", "date_start", "date_end"]
            if file_headers != field_list_to_import:
                raise UserError(
                    _("File headers must match the sample file's headers. \nExpected: %s \nFound: %s") % (field_list_to_import, file_headers))
            # for line in model_rec.item_ids:
            #     if line.product_tmpl_id.barcode not in codes:
            #         line.unlink()
        elif not context.get("sale_direct") and model_name == "rma.ret.mer.auth" and model_rec.rma_type in ("direct", "supplier"):
            field_list_to_import = ["isbn", "source_location_id", "source_parent_id", "destination_location_id",
                                    "destination_parent_id", "order_quantity", "delivered_quantity", "refund_qty",
                                    "price_unit", "tax_id", "refund_price"]
            codes_set = set(codes)
            if model_rec.rma_type == "direct":
                for line in model_rec.rma_direct_lines_ids:
                    if line.product_id.barcode not in codes_set and model_rec.state in ("new", "verification"):
                        line.unlink()
            if model_rec.rma_type == "supplier":
                for line in model_rec.rma_purchase_lines_ids:
                    if line.product_id.barcode not in codes_set and model_rec.state in ("new", "verification"):
                        line.unlink()
        elif model_name == "rma.ret.mer.auth" and model_rec.rma_type == "sale_direct":
            field_list_to_import = ["isbn", "source_location_id", "source_parent_id", "destination_location_id",
                                    "destination_parent_id", "order_quantity", "delivered_quantity", "refund_qty",
                                    "price_unit", "tax_id", "refund_price",
                                    "lco", "mco", "sco", "oco", "tco"]
            codes_set = set(codes)
            for line in model_rec.rma_sale_direct_lines_ids:
                if line.product_id.barcode not in codes_set and model_rec.state in ("new", "verification"):
                    line.unlink()
        elif model_name == "rma.ret.mer.auth" and model_rec.rma_type == "customer":
            field_list_to_import = ["isbn", "source_location_id", "source_parent_id", "destination_location_id",
                                    "destination_parent_id", "order_quantity", "delivered_quantity", "refund_qty",
                                    "price_unit", "tax_id", "refund_price"]
            codes_set = set(codes)
            for line in model_rec.rma_sale_lines_ids:
                if line.product_id.barcode not in codes_set and model_rec.state in ("new", "verification"):
                    line.unlink()
        elif file_headers != field_list_to_import:
            raise UserError(
                _("File headers must match the sample file's headers. \nExpected: %s \nFound: %s") % (field_list_to_import, file_headers))
        if file_headers[-1] == 'bill discount' and model_name == 'purchase.order':
            self.import_po_line_from_export_file(field_list_to_import, sheet)
        # product_cache is now already populated from the pre-check above
        vals_list = []
        code = ''
        list_records = set()
        item_cache = {}
        if model_name == "product.pricelist":
            # Pre-fetch all items for this pricelist to avoid searches in the loop
            existing_items = self.env['product.pricelist.item'].search([('pricelist_id', '=', model_id)])
            for item in existing_items:
                if item.product_tmpl_id.barcode:
                    item_cache[item.product_tmpl_id.barcode] = item
                if item.id:
                    item_cache[str(item.id)] = item

        rma_import_caches = {}
        rma_batch = None
        imported_line_count = 0
        if model_name == "rma.ret.mer.auth":
            rma_import_caches = self._prepare_rma_import_caches(model_rec, context)
            rma_batch = {'create': [], 'updates': []}
        
        if not model_name == "account.move":
            if is_xlsx:
                row_no = 1
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_no += 1
                    passing_dict = dict(zip(field_list_to_import, row))
                    
                    # Validation for empty rows
                    if not passing_dict.get('id') and not passing_dict.get('isbn'):
                         continue

                    code = self.validate_and_get_item_code(passing_dict)
                    if code in list_records and not model_name in ('purchase.order', 'stock.picking') and not (model_name == 'rma.ret.mer.auth' and model_rec.rma_type in ('sale_direct', 'customer', 'supplier')):
                        raise UserError(_(f"Duplicate isbn found: {code}"))
                    else:
                        list_records.add(code)
                    self.create_update_line(
                        passing_dict, model_rec, model_name, vals_list, row_no,
                        product_cache=product_cache, item_cache=item_cache,
                        rma_import_caches=rma_import_caches, rma_batch=rma_batch)
                    if model_name == "rma.ret.mer.auth":
                        imported_line_count += 1
                    # to_be_cancelled_value = passing_dict['to be cancelled'].strip() if passing_dict.get('to be cancelled') else 0
                    to_be_cancelled_value = passing_dict['to be cancelled'] if passing_dict.get(
                        'to be cancelled') is not None and not isinstance(passing_dict.get('to be cancelled'), str) else 0
                    if to_be_cancelled_value != '' and int(to_be_cancelled_value) > 0 and not passing_dict['cancel reason']:
                        title = _("Cancel Reason Required!")
                        message = _(
                            "Data Not Imported! Please check your import file when you add cancel qty, cancel reason is required.")
                        return {
                            'type': 'ir.actions.client',
                            'tag': 'display_notification',
                            'params': {
                                'title': title,
                                'message': message,
                                # 'sticky': False,
                                'type': 'warning',
                            }
                        }
            else: # xlrd
                for row_no in range(sheet.nrows):
                    if row_no != 0:
                        passing_dict = dict(
                            zip(field_list_to_import, sheet.row_values(row_no)))
                    code = self.validate_and_get_item_code(passing_dict)
                    if code in list_records and not model_name in ('purchase.order', 'stock.picking') and not (model_name == 'rma.ret.mer.auth' and model_rec.rma_type in ('sale_direct', 'customer', 'supplier')):
                        raise UserError(_(f"Duplicate isbn found: {code}"))
                    else:
                        list_records.add(code)
                    self.create_update_line(
                        passing_dict, model_rec, model_name, vals_list, row_no,
                        product_cache=product_cache, item_cache=item_cache,
                        rma_import_caches=rma_import_caches, rma_batch=rma_batch)
                    if model_name == "rma.ret.mer.auth":
                        imported_line_count += 1
                    # to_be_cancelled_value = passing_dict['to be cancelled'].strip() if passing_dict.get('to be cancelled') else 0
                    to_be_cancelled_value = passing_dict['to be cancelled'] if passing_dict.get(
                        'to be cancelled') is not None and not isinstance(passing_dict.get('to be cancelled'), str) else 0
                    if to_be_cancelled_value != '' and int(to_be_cancelled_value) > 0 and not passing_dict['cancel reason']:
                        title = _("Cancel Reason Required!")

                        message = _(
                            "Data Not Imported! Please check your import file when you add cancel qty, cancel reason is required.")
                        return {
                            'type': 'ir.actions.client',
                            'tag': 'display_notification',
                            'params': {
                                'title': title,
                                'message': message,
                                # 'sticky': False,
                                'type': 'warning',
                            }
                        }
            if model_name == 'purchase.order':
                po = self.env['purchase.order'].browse(model_rec.id)
                existing_po_line = po.order_line.filtered(lambda l: l.product_id.barcode in [str(val.get("isbn")) for val in vals_list])
                for val in vals_list:
                    isbn = str(val.get("isbn"))
                    if existing_po_line and isbn in (existing_po_line.mapped('product_id.barcode')):
                        for line in po.order_line:
                            if isbn == line.product_id.barcode:
                                line.write(val)
                    else:
                        self.env['purchase.order.line'].create(val)
            
            if model_name == "product.pricelist":
                create_vals = [v for v in vals_list if isinstance(v, dict) and '_is_new' in v]
                for v in create_vals: v.pop('_is_new') # Remove marker
                if create_vals:
                    self.env['product.pricelist.item'].create(create_vals)

            if model_name == "rma.ret.mer.auth" and rma_batch:
                self._flush_rma_batch(model_rec, rma_batch, rma_import_caches)

        if is_xlsx and workbook:
            workbook.close()
        fp.close()
        if os.path.exists(fp.name):
            os.unlink(fp.name)

        if model_name == "rma.ret.mer.auth" and imported_line_count:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Import Successful'),
                    'message': _('Successfully imported %s RMA line(s).') % imported_line_count,
                    'type': 'success',
                    'sticky': False,
                    'next': {'type': 'ir.actions.act_window_close'},
                },
            }
        return True

    def _chunked_in(self, items, chunk_size=2000):
        """Yield chunks of a list for batched IN-domain searches."""
        for idx in range(0, len(items), chunk_size):
            yield items[idx:idx + chunk_size]

    def _build_product_cache(self, codes):
        """Build a product lookup cache from ISBN/codes in chunks."""
        product_cache = {}
        if not codes:
            return product_cache
        unique_codes = list(dict.fromkeys(codes))
        for chunk in self._chunked_in(unique_codes):
            domain = ['|', '|', ('barcode', 'in', chunk), ('default_code', 'in', chunk), ('name', 'in', chunk)]
            for product in self.env['product.product'].search(domain):
                for val in (product.barcode, product.default_code, product.name):
                    if val:
                        product_cache[val] = product
        remaining = [c for c in unique_codes if c not in product_cache]
        for chunk in self._chunked_in(remaining):
            domain = ['|', '|', ('barcode', 'in', chunk), ('default_code', 'in', chunk), ('name', 'in', chunk)]
            for tmpl in self.env['product.template'].search(domain):
                for val in (tmpl.barcode, tmpl.default_code, tmpl.name):
                    if val:
                        product_cache[val] = tmpl
        return product_cache

    def _prepare_rma_import_caches(self, model_rec, context):
        """Pre-fetch locations, taxes, order lines and existing RMA lines once."""
        caches = {
            'location_cache': {},
            'tax_cache': {},
            'po_line_cache': {},
            'so_line_cache': {},
            'rma_line_cache': {},
            'onchange_done': False,
            'context': context,
        }
        if model_rec.rma_type == 'supplier' and model_rec.purchase_order_id:
            po_line_cache = {}
            po_product_counts = defaultdict(int)
            for line in model_rec.purchase_order_id.order_line:
                po_product_counts[line.product_id.id] += 1
                po_line_cache.setdefault(line.product_id.id, line)
            caches['po_line_cache'] = po_line_cache
            caches['po_duplicate_products'] = {
                pid for pid, count in po_product_counts.items() if count > 1
            }
        if model_rec.rma_type == 'customer' and model_rec.sale_order_id:
            so_line_cache = {}
            so_product_counts = defaultdict(int)
            for line in model_rec.sale_order_id.order_line:
                so_product_counts[line.product_id.id] += 1
                so_line_cache.setdefault(line.product_id.id, line)
            caches['so_line_cache'] = so_line_cache
            caches['so_duplicate_products'] = {
                pid for pid, count in so_product_counts.items() if count > 1
            }
        if model_rec.rma_type == 'direct':
            caches['rma_line_cache'] = {
                (line.product_id.id, line.source_location_id.id): line
                for line in model_rec.rma_direct_lines_ids
            }
        elif model_rec.rma_type == 'supplier':
            caches['rma_line_cache'] = {
                (line.product_id.id, line.source_location_id.id): line
                for line in model_rec.rma_purchase_lines_ids
            }
        elif model_rec.rma_type == 'sale_direct':
            caches['rma_line_cache'] = {
                (line.product_id.id, line.destination_location_id.id): line
                for line in model_rec.rma_sale_direct_lines_ids
            }
        elif model_rec.rma_type == 'customer':
            caches['rma_line_cache'] = {
                (line.product_id.id, line.destination_location_id.id): line
                for line in model_rec.rma_sale_lines_ids
            }
        return caches

    def _resolve_rma_locations(self, values, model_rec, location_cache):
        """Resolve source/destination locations once per unique combination."""
        cache_key = (
            values.get('source_location_id'),
            values.get('source_parent_id'),
            values.get('destination_location_id'),
            values.get('destination_parent_id'),
        )
        if cache_key in location_cache:
            return location_cache[cache_key]

        company_id = model_rec.company_id.id
        company_domain = ['|', ('company_id', '=', company_id), ('company_id', '=', False)]
        source_parent = self.env['stock.location'].search(
            [('name', 'ilike', values.get('source_parent_id'))] + company_domain,
            order='company_id desc', limit=1,
        )
        destination_parent = self.env['stock.location'].search(
            [('name', 'ilike', values.get('destination_parent_id'))] + company_domain,
            order='company_id desc', limit=1,
        )
        source_location = self.env['stock.location'].search(
            [('name', 'ilike', values.get('source_location_id'))] + company_domain + [
                '|', ('location_id', '=', source_parent.id), ('location_id', '=', False),
            ],
            order='company_id desc', limit=1,
        )
        dest_location = self.env['stock.location'].search(
            [('name', 'ilike', values.get('destination_location_id'))] + company_domain + [
                '|', ('location_id', '=', destination_parent.id), ('location_id', '=', False),
            ],
            order='company_id desc', limit=1,
        )
        if not source_location or not dest_location:
            raise UserError(_(
                "No location found for source '%s' or destination '%s'.",
                values.get('source_location_id'),
                values.get('destination_location_id'),
            ))
        location_cache[cache_key] = (source_location, dest_location)
        return source_location, dest_location

    def _get_tax_ids_from_cache(self, tax_value, tax_type, tax_cache):
        """Resolve tax ids from a comma-separated tax string using cache."""
        if not tax_value:
            return []
        tax_names = [tax.strip() for tax in str(tax_value).split(',') if tax.strip()]
        tax_ids = []
        for tax_name in tax_names:
            cache_key = (tax_name.lower(), tax_type)
            if cache_key not in tax_cache:
                tax_rec = self.env['account.tax'].search([
                    ('name', 'ilike', tax_name),
                    ('type_tax_use', '=', tax_type),
                ], limit=1)
                tax_cache[cache_key] = tax_rec.ids
            tax_ids.extend(tax_cache[cache_key])
        return list(dict.fromkeys(tax_ids))

    def _flush_rma_batch(self, model_rec, rma_batch, rma_import_caches):
        """Batch-create and batch-update RMA lines to avoid per-row ORM overhead."""
        ctx = dict(self.env.context, mail_notrack=True, tracking_disable=True)
        rma_type = model_rec.rma_type
        if rma_type == 'direct':
            line_model = self.env['rma.direct.lines'].with_context(ctx)
        elif rma_type == 'supplier':
            line_model = self.env['rma.purchase.lines'].with_context(ctx)
        elif rma_type == 'sale_direct':
            line_model = self.env['rma.sale.direct.lines'].with_context(ctx)
        elif rma_type == 'customer':
            line_model = self.env['rma.sale.lines'].with_context(ctx)
        else:
            return

        for line_rec, vals in rma_batch['updates']:
            line_rec.with_context(ctx).write(vals)

        create_vals = rma_batch['create']
        for chunk_start in range(0, len(create_vals), 500):
            line_model.create(create_vals[chunk_start:chunk_start + 500])

    def is_number(self, data):
        if data is None:
            return False
        try:
            float(data)
            return True
        except (ValueError, TypeError):
            return False

    def validate_and_get_item_code(self, value):
        """Method to get isbn(barcode) code of product"""
        item_code = value.get("isbn")
        if not item_code:
            raise UserError(_("Isbn is required."))
        item_code = str(item_code).strip()
        if len(item_code) <= 0:
            raise UserError(_("Isbn is required."))
        if self.is_number(item_code):
            return int(float(item_code))
        return item_code

    def validate_and_get_done_qty(self, value):
        """Method to update get done qty"""
        done = value.get("done")
        # remove this validation for task 02969
        # if not done:
        #     raise UserError(_("Done is required."))
        if done is None:
            return 0.0
        done = str(done).strip()
        if len(done) <= 0:
            return 0.0
        if self.is_number(done):
            return float(done)
        return 0.0

    def create_update_line(self, values, model_rec, model_name, vals_list=[], row_no=None,
                          product_cache=None, item_cache=None, rma_import_caches=None, rma_batch=None):
        """Method to update the done qty of picking order"""
        if product_cache is None: product_cache = {}
        if item_cache is None: item_cache = {}
        if rma_import_caches is None: rma_import_caches = {}
        prod_tmpl = values.get("product_tmpl")
        code = self.validate_and_get_item_code(values)
        cache_code = str(code)
        
        # --- CACHED PRODUCT SEARCH ---
        if cache_code not in product_cache:
            domain = [
                '|', '|',
                ('barcode', '=', code),
                ('default_code', '=', code),
                ('name', '=', code)
            ]
            if model_name == "product.pricelist":
                product = self.env["product.template"].search(domain, limit=1)
                if not product:
                    product = self.env["product.product"].search(domain, limit=1)
            else:
                product = self.env["product.product"].search(domain, limit=1)
            product_cache[cache_code] = product
        
        product = product_cache[cache_code]
        if not product:
            raise UserError(_(f"No product Found with isbn: {code}"))
        po_line_obj = self.env["purchase.order.line"]
        po_request_line_obj = self.env["purchase.request.line"]
        rma_line_obj = self.env["rma.direct.lines"]
        rma_sale_direct_obj = self.env["rma.sale.direct.lines"]
        rma_sale_line_obj = self.env["rma.sale.lines"]
        rma_purchase_line_obj = self.env["rma.purchase.lines"]

        context = self.env.context

        if model_name == "purchase.order" and len(values) > 4:
            cso = values.get("customer sales order")
            customer_name = values.get("customer name")
            sale_ref = values.get("sales ref- sor odoo")
            related_so = self.env['sale.order'].search([('name', '=', sale_ref)], limit=1)
            if not related_so and sale_ref:
                raise UserError("PLease Correct the value of 'Sales Ref- SOR Odoo' {} as the inputted value in this field is not correct. Please check the import file again of Product Line {}".format(sale_ref, code))
            elif not sale_ref:
                raise UserError("Please set the value in the Column 'Sales Ref- SOR Odoo' as it's a required column.")
            po_line_rec = po_line_obj.search(
                [
                    ("product_id.barcode", "=", code),
                    ("order_id", "=", model_rec and model_rec.id),
                ], limit=1
            )
            po_line_id = values.get("id") or False
            if po_line_id:
                po_line_rec = po_line_obj.search([('id', '=', int(po_line_id)), ("product_id.barcode", "=", code),
                                                  ("order_id", "=", model_rec and model_rec.id)])

            rfq_val = values.get("rfq quantity")
            cancel_val = values.get("to be cancelled")
            to_be_received_val = values.get("to be received qty")

            rfq_qty = float(rfq_val) if self.is_number(rfq_val) else 0
            cancel_qty = float(cancel_val) if self.is_number(cancel_val) else 0
            to_be_received_qty = float(to_be_received_val) if self.is_number(to_be_received_val) else 0

            vals = {
                    "po_qty": rfq_qty - cancel_qty,
                    "rfq_qty": rfq_qty,
                    "cancel_qty": cancel_qty,
                    "isbn": code,
                    "po_list_price": float(values.get("po list price")) if self.is_number(values.get("po list price")) else 0,
                    "po_target_price": values.get("po target price") or '',
                    "list_price": float(values.get("bill list price")) if self.is_number(values.get("bill list price")) else 0,
                    "po_discount": float(values.get("po discount")) if self.is_number(values.get("po discount")) else 0,
                    "discount": float(values.get("bill discount")) if self.is_number(values.get("bill discount")) else 0,
                    "customer_sales_order": cso or '',
                    "customer_name": customer_name or '',
                    "related_so": related_so.id if related_so else False,
                    "remarks": values.get("remarks") or '',
                    "cancel_reason": values.get("cancel reason") or '',
                    }
            if model_rec.state in ("draft", "sent"):
                if not po_line_rec:
                    vals.update({
                        "product_id": product and product.id,
                        "order_id": model_rec and model_rec.id,
                        "rfq_qty": rfq_qty,
                        # if po is not confirm then to_be_received_qty == po_qty and here po_qty calculate based on the rfq_qty and to_be_cancelled_qty
                        "to_be_received_qty": rfq_qty - cancel_qty,
                        "product_qty": to_be_received_qty,})
                    vals_list.append(vals)
                else:
                    po_line_rec.with_context(update_import=True).write(vals)
                    po_line_rec.to_be_received_qty = to_be_received_qty
            elif model_rec and model_rec.state == "purchase" and po_line_rec:
                po_line_rec.with_context(update_import=True).write(vals)
                po_line_rec.to_be_received_qty = to_be_received_qty
            # Have to uncomment this code on 16th Jan 2024
            # if po_line_rec.po_qty < po_line_rec.product_qty:
            #     raise ValidationError(_('You can not set Vendor Bill Qty more then PO Quantity'))

            return vals_list

        if model_name == "purchase.request":
            po_request_line = po_request_line_obj.create(
                {
                    "product_id": product and product.id,
                    "product_qty": values.get("qty"),
                    "estimated_cost": values.get("estimated cost"),
                    "request_id": model_rec and model_rec.id,
                }
            )
            po_request_line.onchange_product_id()
        # PRICELIST
        if model_name == "product.pricelist":
            pricelist_item_obj = self.env['product.pricelist.item']
            item_id = values.get('id')
            pricelist_item_rec = False
            
            # --- CACHED ITEM SEARCH ---
            if item_id and str(int(float(item_id))) in item_cache:
                pricelist_item_rec = item_cache[str(int(float(item_id)))]
            elif code in item_cache:
                pricelist_item_rec = item_cache[code]

            # Auto-detect compute_price if missing
            compute_price = str(values.get("compute_price") or '').lower().strip()
            fixed_price = float(values.get("fixed_price") or 0.0)
            percent_price = float(values.get("percent_price") or 0.0)
            
            if not compute_price:
                if fixed_price > 0:
                    compute_price = 'fixed'
                elif percent_price > 0:
                    compute_price = 'percentage'
                else:
                    compute_price = 'fixed' # Default

            # Auto-detect applied_on if missing
            applied_on = str(values.get("applied_on") or '').lower().strip()
            if not applied_on:
                applied_on = '1_product' # Default for ISBN import

            vals = {
                "product_tmpl_id": product and product.id,
                "fixed_price": fixed_price,
                "percent_price": percent_price,
                "compute_price": compute_price,
                "applied_on": applied_on,
                "min_quantity": values.get("min_quantity") or 0.0,
                "date_start": values.get("date_start"),
                "date_end": values.get("date_end"),
                "pricelist_id": model_rec and model_rec.id,
            }
            if pricelist_item_rec:
                # Use Direct SQL for updates to bypass Odoo's N+1 write overhead
                self.env.cr.execute("""
                    UPDATE product_pricelist_item 
                    SET fixed_price = %s, 
                        percent_price = %s, 
                        compute_price = %s, 
                        applied_on = %s, 
                        min_quantity = %s, 
                        date_start = %s, 
                        date_end = %s
                    WHERE id = %s
                """, (
                    vals['fixed_price'], 
                    vals['percent_price'], 
                    vals['compute_price'], 
                    vals['applied_on'], 
                    vals['min_quantity'], 
                    vals.get('date_start') or None, 
                    vals.get('date_end') or None, 
                    pricelist_item_rec.id
                ))
                # Mark model for cache invalidation at the end of the transaction
                model_rec.invalidate_recordset(['item_ids'])
            else:
                # Add to list for batch create
                vals.update({'_is_new': True})
                vals_list.append(vals)

        if model_name == "stock.picking":
            id_val = values.get("id")
            if not id_val:
                raise UserError(_("ID is required for stock picking import."))
            id = int(float(id_val))
            rec = model_rec.move_ids.search(
                [
                    ("id", "=", id),
                    ("product_id.barcode", "=", code),
                    ("picking_id", "=", model_rec.id),
                ]
            )
            if len(rec) == 1:
                qty_to_update = self.validate_and_get_done_qty(values)
                # if rec.product_uom_qty < qty_to_update:
                #     raise UserError(
                #         _("Done qty should not be greater than Demand Qty."))
                rec.write({"quantity": qty_to_update, "cancel_reason": values.get('cancel_reason')})
            if len(rec) > 1:
                done_qty_to_update = self.validate_and_get_done_qty(values)
                for mv in rec:
                    mv.cancel_reason = values.get('cancel_reason')
                    demanded_qty = mv.product_uom_qty
                    if (demanded_qty <= done_qty_to_update):
                        mv.quantity = demanded_qty
                    elif demanded_qty >= done_qty_to_update:
                        mv.quantity = done_qty_to_update
                    done_qty_to_update = done_qty_to_update - mv.quantity
        if model_name == "rma.ret.mer.auth":
            order_quantity = values.get("order_quantity")
            delivered_quantity = values.get("delivered_quantity")
            refund_qty = values.get("refund_qty")
            price_unit = values.get("price_unit")
            refund_price = values.get("refund_price")
            landed_cost = values.get("lco")
            subtotal_cost = values.get("tco")
            source_location, dest_location = self._resolve_rma_locations(
                values, model_rec, rma_import_caches.setdefault('location_cache', {}))
            if order_quantity and delivered_quantity and refund_qty:
                if refund_qty > delivered_quantity or refund_qty > order_quantity:
                    raise UserError(
                        _("Return qty should not exceed to order qty or delivered qty!"))
            if not rma_import_caches.get('onchange_done'):
                model_rec._onchange_sale_purchase_type()
                rma_import_caches['onchange_done'] = True
            tax_list = []
            # Direct PO Import
            if model_rec.rma_type in ("direct", "supplier") and not context.get("sale_direct"):
                tax_list = self._get_tax_ids_from_cache(
                    values.get("tax_id"), 'purchase', rma_import_caches.setdefault('tax_cache', {}))
                if model_rec.rma_type == 'supplier' and (
                        not order_quantity or not delivered_quantity or not price_unit or not tax_list):
                    if product.id in rma_import_caches.get('po_duplicate_products', set()):
                        raise UserError(
                            _("Duplicate product found for CSO Reference %s, can not import it, kindly create RMA manually")
                            % model_rec.name)
                    po_line_rec = rma_import_caches.get('po_line_cache', {}).get(product.id)
                    if po_line_rec:
                        if not order_quantity:
                            order_quantity = po_line_rec.product_qty
                        if not delivered_quantity:
                            delivered_quantity = po_line_rec.qty_received
                        if not price_unit:
                            price_unit = po_line_rec.price_unit
                        if not tax_list:
                            tax_list = po_line_rec.tax_ids.ids
                if not order_quantity and refund_qty:
                    order_quantity = refund_qty
                if not delivered_quantity and refund_qty:
                    delivered_quantity = refund_qty
                rma_po_vals = {
                    "product_id": product and product.id,
                    "source_location_id": source_location.id,
                    "destination_location_id": dest_location.id,
                    "order_quantity": order_quantity,
                    "delivered_quantity": delivered_quantity,
                    "refund_qty": refund_qty,
                    "price_unit": price_unit,
                    "tax_id": [(6, 0, tax_list)],
                    "refund_price": refund_price,
                    "rma_id": model_rec.id,
                }
                line_key = (product.id, source_location.id)
                existing_line = rma_import_caches.get('rma_line_cache', {}).get(line_key)
                if rma_batch is not None:
                    if existing_line:
                        rma_batch['updates'].append((existing_line, rma_po_vals))
                    else:
                        rma_batch['create'].append(rma_po_vals)
                        rma_import_caches.setdefault('rma_line_cache', {})[line_key] = True
                elif existing_line and existing_line is not True:
                    existing_line.write(rma_po_vals)
                elif model_rec.rma_type == "direct":
                    new_line = rma_line_obj.create(rma_po_vals)
                    rma_import_caches.setdefault('rma_line_cache', {})[line_key] = new_line
                elif model_rec.rma_type == "supplier":
                    rma_purchase_line_obj.create(rma_po_vals)
            # Direct SO Import
            elif model_rec.rma_type in ("sale_direct", "customer") or context.get("sale_direct"):
                tax_list = self._get_tax_ids_from_cache(
                    values.get("tax_id"), 'sale', rma_import_caches.setdefault('tax_cache', {}))
                if model_rec.rma_type == 'customer' and (
                        not order_quantity or not delivered_quantity or not price_unit
                        or not tax_list or not landed_cost):
                    if product.id in rma_import_caches.get('so_duplicate_products', set()):
                        raise UserError(
                            _("Duplicate product found for CSO Reference %s, can not import it, kindly create RMA manually")
                            % model_rec.name)
                    so_line_rec = rma_import_caches.get('so_line_cache', {}).get(product.id)
                    if so_line_rec:
                        if not order_quantity:
                            order_quantity = so_line_rec.product_uom_qty
                        if not delivered_quantity:
                            delivered_quantity = so_line_rec.qty_delivered
                        if not price_unit:
                            price_unit = so_line_rec.price_unit
                        if not tax_list:
                            tax_list = so_line_rec.tax_ids.ids
                        if not landed_cost:
                            landed_cost = so_line_rec.landed_cost
                if not order_quantity and refund_qty:
                    order_quantity = refund_qty
                if not delivered_quantity and refund_qty:
                    delivered_quantity = refund_qty
                if not subtotal_cost:
                    mco_val = values.get("mco")
                    sco_val = values.get("sco")
                    oco_val = values.get("oco")
                    mco = float(mco_val) if self.is_number(mco_val) else 0.0
                    sco = float(sco_val) if self.is_number(sco_val) else 0.0
                    oco = float(oco_val) if self.is_number(oco_val) else 0.0
                    subtotal_cost = sum([landed_cost or 0, mco, sco, oco])
                rma_so_vals = {
                    "product_id": product and product.id,
                    "source_location_id": source_location.id,
                    "destination_location_id": dest_location.id,
                    "order_quantity": order_quantity,
                    "delivered_quantity": delivered_quantity,
                    "refund_qty": refund_qty,
                    "price_unit": price_unit,
                    "tax_id": [(6, 0, tax_list)],
                    "refund_price": refund_price,
                    "rma_id": model_rec.id,
                    "landed_cost": landed_cost,
                    "marketplace_cost": values.get("mco"),
                    "shipping_cost": values.get("sco"),
                    "other_cost": values.get("oco"),
                    "subtotal_cost": subtotal_cost,
                }
                line_key = (product.id, dest_location.id)
                existing_line = rma_import_caches.get('rma_line_cache', {}).get(line_key)
                if rma_batch is not None:
                    if existing_line and existing_line is not True:
                        rma_batch['updates'].append((existing_line, rma_so_vals))
                    else:
                        rma_batch['create'].append(rma_so_vals)
                        rma_import_caches.setdefault('rma_line_cache', {})[line_key] = True
                elif model_rec.rma_type == "sale_direct":
                    rma_sale_direct_obj.create(rma_so_vals)
                elif existing_line and existing_line is not True:
                    existing_line.write(rma_so_vals)
                elif model_rec.rma_type == "customer":
                    rma_sale_line_obj.create(rma_so_vals)



    def create_attachment(self, file_name):
        """
        Delete file created in tmp dir, Delete old attachment and create attachment for download report
        :param file_name: file name string
        :return: attachment ir.attachment object
        """
        ir_attachment_obj = self.env['ir.attachment']
        # Read File data
        with open(f'/tmp/{file_name}', "rb+") as file:
            file_data = base64.encodebytes(file.read())
            file.close()

        # Remove tmp file
        os.remove(f'/tmp/{file_name}')
        # Delete Old Attachment
        attachments = ir_attachment_obj.search([('name', '=ilike', file_name),
                                                ('res_model', '=', 'purchase.order')])
        attachments and attachments.unlink()

        return ir_attachment_obj.create({
            'name': file_name,
            'datas': file_data,
            'res_model': 'purchase.order',
            'type': 'binary'
        })

    def export_data(self):
        report_file_name = ''
        """Method to export purchase order lines with difference for compare actions"""
        if self.env.context.get('active_model') == 'purchase.order':
            report_file_name = self.prepare_export_excel_data()
        if self.env.context.get('active_model') == 'stock.picking':
            report_file_name = self.prepare_picking_export_excel_data()
        if self.env.context.get('active_model') == 'account.move':
            report_file_name = self.prepare_account_move_export_excel_data()
        if self.env.context.get('active_model') == 'product.pricelist':
            report_file_name = self.prepare_pricelist_export_excel_data()
        # Create Attachment
        attachment = self.create_attachment(report_file_name)
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'download',
        }

    def prepare_export_excel_data(self):
        active_id = self.env.context.get("active_id")
        po_id = self.env["purchase.order"].browse(active_id)
        po_lines = po_id.order_line
        file_name = 'po_lines_data.xlsx'
        workbook = xlsxwriter.Workbook(f'/tmp/{file_name}')
        worksheet = workbook.add_worksheet("POLinesData")
        worksheet.protect()
        worksheet.set_landscape()
        worksheet.fit_to_pages(1, 0)
        worksheet.set_zoom(80)
        worksheet.set_column(1, 1, 20)
        worksheet.set_column(2, 2, 20)
        worksheet.set_column(3, 3, 15)
        worksheet.set_column(4, 4, 15)
        worksheet.set_column(6, 6, 20)

        qty_data = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'locked': False})
        qty_data_lock = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
        header_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'bold': True, 'border': 1})

        row, col = 0, 0
        worksheet.set_row(1, 30)
        worksheet.write(row, col, 'Id', header_format)
        col += 1
        worksheet.write(row, col, 'Isbn', header_format)
        col += 1
        worksheet.write(row, col, 'RFQ Quantity', header_format)
        col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'To Be Cancelled', header_format)
        col += 1
        worksheet.set_column(col, col, 40)
        worksheet.write(row, col, 'To Be Received QTY', header_format)
        col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'Po List Price', header_format)
        col += 1
        worksheet.write(row, col, 'Po Discount', header_format)
        col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'Po Target Price', header_format)
        col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'Bill List Price', header_format)
        col += 1
        worksheet.set_column(col, col, 20)
        worksheet.write(row, col, 'Bill Discount', header_format)
        # col += 1
        # worksheet.set_column(col, col, 20)
        # worksheet.write(row, col, 'Customer Sales Order', header_format)
        # col += 1
        # worksheet.set_column(col, col, 20)
        # worksheet.write(row, col, 'Customer Name', header_format)
        col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'Sales Ref- SOR Odoo', header_format)
        col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'Remarks', header_format)
        col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'Cancel Reason', header_format)
        col += 1
        row = 1

        for line in po_lines:
            col = 0
            diff_amt = line.price_unit - line.po_price
            worksheet.write(row, col, line.id or '', qty_data_lock)
            col += 1
            worksheet.write(row, col, line.product_id.barcode or '', qty_data_lock)
            col += 1
            worksheet.write(row, col, line.rfq_qty or 0, qty_data_lock)
            col += 1
            worksheet.write(row, col, line.cancel_qty or 0, qty_data)
            col += 1
            worksheet.write(row, col, line.to_be_received_qty or 0, qty_data)
            col += 1
            worksheet.write(row, col, line.po_list_price or 0, qty_data_lock)
            col += 1
            worksheet.write(row, col, line.po_discount or 0, qty_data_lock)
            col += 1
            worksheet.write(row, col, line.po_target_price or 0, qty_data_lock)
            col += 1
            worksheet.write(row, col, line.list_price or 0, qty_data)
            col += 1
            worksheet.write(row, col, line.discount or 0, qty_data)
            col += 1
            # worksheet.write(row, col, line.customer_sales_order or '', qty_data)
            # col += 1
            # worksheet.write(row, col, line.customer_name or '', qty_data)
            # col += 1
            # worksheet.write(row, col, line.sale_reference or '', qty_data)
            worksheet.write(row, col, line.related_so.name if line.related_so else False, qty_data)
            col += 1
            worksheet.write(row, col, line.remarks or '', qty_data)
            col += 1
            cancel_reason_selection = [l[0] for l in line._fields['cancel_reason'].selection]
            worksheet.data_validation(row, 14, row, 14, {'validate': 'list', 'source': cancel_reason_selection})
            worksheet.write(row, col, line.cancel_reason or '', qty_data)
            col += 1
            row += 1

        workbook.close()
        return file_name

    def prepare_pricelist_export_excel_data(self):
        active_id = self.env.context.get("active_id")
        pricelist_id = self.env["product.pricelist"].browse(active_id)
        items = pricelist_id.item_ids
        file_name = 'pricelist_items_data.xlsx'
        workbook = xlsxwriter.Workbook(f'/tmp/{file_name}')
        worksheet = workbook.add_worksheet("PricelistData")
        
        header_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'bold': True, 'border': 1})
        data_format = workbook.add_format({'align': 'center', 'valign': 'vcenter'})
        
        headers = ["id", "isbn", "applied_on", "compute_price", "fixed_price", "percent_price", "min_quantity", "date_start", "date_end"]
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
            worksheet.set_column(col, col, 15)

        row = 1
        for item in items:
            worksheet.write(row, 0, item.id, data_format)
            worksheet.write(row, 1, item.product_tmpl_id.barcode or '', data_format)
            worksheet.write(row, 2, item.applied_on or '', data_format)
            worksheet.write(row, 3, item.compute_price or '', data_format)
            worksheet.write(row, 4, item.fixed_price or 0.0, data_format)
            worksheet.write(row, 5, item.percent_price or 0.0, data_format)
            worksheet.write(row, 6, item.min_quantity or 0.0, data_format)
            worksheet.write(row, 7, str(item.date_start) if item.date_start else '', data_format)
            worksheet.write(row, 8, str(item.date_end) if item.date_end else '', data_format)
            row += 1

        workbook.close()
        return file_name

    def prepare_picking_export_excel_data(self):
        active_id = self.env.context.get("active_id")
        picking_id = self.env["stock.picking"].browse(active_id)
        picking_lines = picking_id.move_ids
        file_name = 'picking_lines_data.xlsx'
        workbook = xlsxwriter.Workbook(f'/tmp/{file_name}')
        worksheet = workbook.add_worksheet("PickingLinesData")
        worksheet.set_landscape()
        worksheet.protect()
        worksheet.fit_to_pages(1, 0)
        worksheet.set_zoom(80)
        worksheet.set_column(1, 1, 20)
        worksheet.set_column(2, 2, 20)
        worksheet.set_column(3, 3, 15)
        worksheet.set_column(4, 4, 15)
        worksheet.set_column(6, 6, 20)

        qty_data = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'locked': False})
        qty_data_lock = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
        header_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'bold': True, 'border': 1})

        row, col = 0, 0
        worksheet.set_row(1, 30)
        worksheet.write(row, col, 'id', header_format)
        col += 1
        worksheet.write(row, col, 'isbn', header_format)
        col += 1
        worksheet.write(row, col, 'bundle_isbn', header_format)
        col += 1
        worksheet.write(row, col, 'demand', header_format)
        col += 1
        worksheet.write(row, col, 'done', header_format)
        col += 1
        worksheet.write(row, col, 'cancel_reason', header_format)
        col += 1

        row = 1

        for line in picking_lines:
            col = 0
            worksheet.write(row, col, line.id or '',)
            col += 1
            worksheet.write(row, col, line.product_id.barcode or '', qty_data)
            col += 1
            worksheet.write(row, col, line.bom_line_id.bom_id.product_tmpl_id.barcode or '', qty_data)
            col += 1
            worksheet.write(row, col, line.product_uom_qty or 0, qty_data)
            col += 1
            worksheet.write(row, col, line.quantity or 0, qty_data)
            col += 1
            cancel_reason_selection = [l[0] for l in line._fields['cancel_reason'].selection]
            worksheet.data_validation(row, 3, row, 3, {'validate': 'list', 'source': cancel_reason_selection})
            worksheet.write(row, col, line.cancel_reason or '', qty_data)
            col += 1
            row += 1

        workbook.close()
        return file_name

    def prepare_account_move_export_excel_data(self):
        active_id = self.env.context.get("active_id")
        account_move_id = self.env["account.move"].browse(active_id)
        move_lines = account_move_id.invoice_line_ids
        if account_move_id.move_type == "entry":
            file_name = 'journal_lines.xlsx'
        elif account_move_id.move_type == "in_invoice":
            file_name = 'bill_lines_data.xlsx'
        else:
            file_name = 'invoice_lines_data.xlsx'
        workbook = xlsxwriter.Workbook(f'/tmp/{file_name}')
        worksheet = workbook.add_worksheet("LinesData")
        worksheet.set_landscape()
        worksheet.fit_to_pages(1, 0)
        worksheet.set_zoom(80)
        worksheet.set_column(1, 1, 20)
        worksheet.set_column(2, 2, 20)
        worksheet.set_column(3, 3, 15)
        worksheet.set_column(4, 4, 15)
        worksheet.set_column(6, 6, 20)
        qty_data = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'locked': False})
        qty_data_lock = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
        header_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'bold': True, 'border': 1})
        row, col = 0, 0
        if account_move_id.move_type == 'entry':
            worksheet.set_row(1, 30)
            worksheet.write(row, col, 'id', header_format)
            col += 1
            worksheet.write(row, col, 'account_id', header_format)
            col += 1
            worksheet.write(row, col, 'partner_id', header_format)
            col += 1
            worksheet.set_column(col, col, 45)
            worksheet.write(row, col, 'label', header_format)
            col += 1
            worksheet.set_column(col, col, 35)
            worksheet.write(row, col, 'analytic_distribution', header_format)
            col += 1
            worksheet.write(row, col, 'debit', header_format)
            col += 1
            worksheet.write(row, col, 'credit', header_format)
            col += 1
            row = 1
            for line in move_lines:
                col = 0
                worksheet.write(row, col, line.id or '', qty_data_lock)
                col += 1
                worksheet.write(row, col, line.account_id.display_name or '', qty_data)
                col += 1
                worksheet.write(row, col, line.partner_id.display_name or '', qty_data)
                col += 1
                worksheet.write(row, col, line.name or '', qty_data)
                col += 1
                analytic_names = ''
                if line.analytic_distribution:
                    analytic_account_ids = [int(aid) for key in line.analytic_distribution for aid in key.split(',') if aid.isdigit()]
                    if analytic_account_ids:
                        analytic_names = ', '.join(self.env['account.analytic.account'].browse(analytic_account_ids).exists().mapped('display_name'))
                worksheet.write(row, col, analytic_names, qty_data)
                col += 1
                worksheet.write(row, col, line.debit or 0, qty_data)
                col += 1
                worksheet.write(row, col, line.credit or 0, qty_data)
                col += 1
                row += 1
        else:
            worksheet.set_row(1, 30)
            worksheet.write(row, col, 'id', header_format)
            col += 1
            worksheet.write(row, col, 'isbn', header_format)
            col += 1
            worksheet.write(row, col, 'description', header_format)
            col += 1
            worksheet.write(row, col, 'quantity', header_format)
            col += 1
            worksheet.write(row, col, 'list_price', header_format)
            col += 1
            worksheet.set_column(col, col, 40)
            worksheet.write(row, col, 'discount', header_format)
            col += 1
            worksheet.set_column(col, col, 15)
            worksheet.write(row, col, 'tax_ids', header_format)
            col += 1
            worksheet.write(row, col, 'account_id', header_format)
            col += 1
            worksheet.write(row, col, 'analytic_distribution', header_format)
            col += 1
            worksheet.set_column(col, col, 15)
            row = 1
            for line in move_lines:
                col = 0
                worksheet.write(row, col, line.id or '', qty_data_lock)
                col += 1
                worksheet.write(row, col, line.product_id.barcode or '', qty_data)
                col += 1
                worksheet.write(row, col, line.name or '', qty_data)
                col += 1
                worksheet.write(row, col, line.quantity or 0, qty_data)
                col += 1
                worksheet.write(row, col, line.price_unit or 0, qty_data)
                col += 1
                worksheet.write(row, col, line.discount or 0, qty_data)
                col += 1
                tax_tags = ', '.join(line.tax_ids.mapped('name'))
                worksheet.write(row, col, tax_tags or '', qty_data)
                col += 1
                worksheet.write(row, col, line.account_id.display_name or '', qty_data)
                col += 1
                analytic_names = ''
                if line.analytic_distribution:
                    analytic_account_ids = [int(aid) for key in line.analytic_distribution for aid in key.split(',') if aid.isdigit()]
                    if analytic_account_ids:
                        analytic_names = ', '.join(self.env['account.analytic.account'].browse(analytic_account_ids).exists().mapped('display_name'))
                worksheet.write(row, col, analytic_names, qty_data)
                col += 1
                row += 1
            workbook.close()
        return file_name

    def export_rma_lines(self):
        """Method to export RMA lines for any returns to Excel"""
        active_id = self.env.context.get("active_id")
        rma_id = self.env["rma.ret.mer.auth"].browse(active_id)
        
        file_name = f'rma_sales_lines_{rma_id.name or "data"}.xlsx'
        workbook = xlsxwriter.Workbook(f'/tmp/{file_name}')
        worksheet = workbook.add_worksheet("RMALinesData")
        worksheet.protect()
        worksheet.set_landscape()
        worksheet.fit_to_pages(1, 0)
        worksheet.set_zoom(80)
        
        # Set column widths
        worksheet.set_column(0, 0, 15)  # ISBN
        worksheet.set_column(1, 1, 20)  # Source Location
        worksheet.set_column(2, 2, 20)  # Source Parent
        worksheet.set_column(3, 3, 20)  # Dest Location
        worksheet.set_column(4, 4, 20)  # Dest Parent
        worksheet.set_column(5, 7, 15)  # Qty columns
        worksheet.set_column(8, 8, 15)  # Price Unit
        worksheet.set_column(9, 9, 20)  # Taxes
        worksheet.set_column(10, 15, 15) # Costs

        # Formats
        qty_data = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
        qty_data_unlock = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'locked': False})
        header_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'bold': True, 'border': 1})

        # Base fields for customer RMA
        headers = ["isbn", "source_location_id", "source_parent_id", "destination_location_id",
                   "destination_parent_id", "order_quantity", "delivered_quantity", "refund_qty",
                   "price_unit", "tax_id", "refund_price"]
        
        if rma_id.rma_type == 'sale_direct' or self.env.context.get("sale_direct"):
             headers.extend(["lco", "mco", "sco", "oco", "tco"])

        # Write Headers
        row, col = 0, 0
        worksheet.set_row(0, 30)
        for header in headers:
            worksheet.write(row, col, header, header_format)
            col += 1
            
        row = 1
        
        # Determine which lines to iterate over based on RMA type
        lines_to_export = []
        if rma_id.rma_type == "customer":
            lines_to_export = rma_id.rma_sale_lines_ids
        elif rma_id.rma_type == "sale_direct":
             lines_to_export = rma_id.rma_sale_direct_lines_ids
        elif rma_id.rma_type == "supplier":
             lines_to_export = rma_id.rma_purchase_lines_ids
        elif rma_id.rma_type == "direct":
             lines_to_export = rma_id.rma_direct_lines_ids

        for line in lines_to_export:
            col = 0
            worksheet.write(row, col, line.product_id.barcode or '', qty_data)
            col += 1
            worksheet.write(row, col, line.source_location_id.name or '', qty_data)
            col += 1
            worksheet.write(row, col, line.source_location_id.location_id.name if line.source_location_id.location_id else '', qty_data)
            col += 1
            worksheet.write(row, col, line.destination_location_id.name or '', qty_data)
            col += 1
            worksheet.write(row, col, line.destination_location_id.location_id.name if line.destination_location_id.location_id else '', qty_data)
            col += 1
            worksheet.write(row, col, line.order_quantity or 0, qty_data)
            col += 1
            worksheet.write(row, col, line.delivered_quantity or 0, qty_data)
            col += 1
            worksheet.write(row, col, line.refund_qty or 0, qty_data_unlock)
            col += 1
            worksheet.write(row, col, line.price_unit or 0, qty_data)
            col += 1
            taxes = ', '.join(line.tax_id.mapped('name'))
            worksheet.write(row, col, taxes or '', qty_data)
            col += 1
            worksheet.write(row, col, line.refund_price or 0, qty_data)
            col += 1
            
            if rma_id.rma_type == 'sale_direct' or self.env.context.get("sale_direct"):
                worksheet.write(row, col, line.landed_cost or 0, qty_data)
                col += 1
                worksheet.write(row, col, line.marketplace_cost or 0, qty_data)
                col += 1
                worksheet.write(row, col, line.shipping_cost or 0, qty_data)
                col += 1
                worksheet.write(row, col, line.other_cost or 0, qty_data)
                col += 1
                worksheet.write(row, col, line.subtotal_cost or 0, qty_data)
                col += 1

            row += 1

        workbook.close()
        
        # Create Attachment
        attachment = self.create_attachment(file_name)
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'download',
        }
    def export_data_update(self):
        report_file_name = ''
        """Method to export purchase order lines with difference for compare actions"""
        if self.env.context.get('active_model') == 'purchase.order':
            report_file_name = self.prepare_export_excel_data_update()
        # Create Attachment
        attachment = self.create_attachment(report_file_name)
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'download',
        }

    def prepare_export_excel_data_update(self):
        active_id = self.env.context.get("active_id")
        po_id = self.env["purchase.order"].browse(active_id)
        po_lines = po_id.order_line
        file_name = 'po_lines_data.xlsx'
        workbook = xlsxwriter.Workbook(f'/tmp/{file_name}')
        worksheet = workbook.add_worksheet("POLinesData")
        worksheet.protect()
        worksheet.set_landscape()
        worksheet.fit_to_pages(1, 0)
        worksheet.set_zoom(80)
        worksheet.set_column(1, 1, 20)
        worksheet.set_column(2, 2, 20)
        worksheet.set_column(3, 3, 15)
        worksheet.set_column(4, 4, 15)
        worksheet.set_column(6, 6, 20)

        qty_data = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'locked': False})
        qty_data_lock = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
        header_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'bold': True, 'border': 1})

        row, col = 0, 0
        worksheet.set_row(1, 30)
        worksheet.write(row, col, 'Id', header_format)
        col += 1
        worksheet.write(row, col, 'Isbn', header_format)
        col += 1
        worksheet.write(row, col, 'RFQ Quantity', header_format)
        col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'To Be Cancelled', header_format)
        col += 1
        worksheet.set_column(col, col, 40)
        worksheet.write(row, col, 'To Be Received QTY', header_format)
        col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'Po List Price', header_format)
        col += 1
        worksheet.write(row, col, 'Po Discount', header_format)
        col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'Po Target Price', header_format)
        col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'Bill List Price', header_format)
        col += 1
        worksheet.set_column(col, col, 20)
        worksheet.write(row, col, 'Bill Discount', header_format)
        col += 1
        worksheet.set_column(col, col, 20)
        # worksheet.write(row, col, 'Customer Sales Order', header_format)
        # col += 1
        # worksheet.set_column(col, col, 20)
        # worksheet.write(row, col, 'Customer Name', header_format)
        # col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'Sales Ref- SOR Odoo', header_format)
        col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'Remarks', header_format)
        col += 1
        worksheet.set_column(col, col, 15)
        worksheet.write(row, col, 'Cancel Reason', header_format)
        col += 1
        row = 1

        for line in po_lines:
            col = 0
            diff_amt = line.price_unit - line.po_price
            worksheet.write(row, col, line.id or '', qty_data)
            col += 1
            worksheet.write(row, col, line.product_id.barcode or '', qty_data)
            col += 1
            worksheet.write(row, col, line.rfq_qty or 0, qty_data_lock)
            col += 1
            worksheet.write(row, col, line.cancel_qty or 0, qty_data)
            col += 1
            worksheet.write(row, col, line.to_be_received_qty or 0, qty_data_lock)
            col += 1
            worksheet.write(row, col, line.po_list_price or 0, qty_data)
            col += 1
            worksheet.write(row, col, line.po_discount or 0, qty_data)
            col += 1
            worksheet.write(row, col, line.po_target_price or 0, qty_data)
            col += 1
            worksheet.write(row, col, line.list_price or 0, qty_data)
            col += 1
            worksheet.write(row, col, line.discount or 0, qty_data)
            col += 1
            # worksheet.write(row, col, line.customer_sales_order or '', qty_data)
            # col += 1
            # worksheet.write(row, col, line.customer_name or '', qty_data)
            # col += 1
            # worksheet.write(row, col, line.sale_reference or '', qty_data)
            worksheet.write(row, col, line.related_so.name if line.related_so else False, qty_data)
            col += 1
            worksheet.write(row, col, line.remarks or '', qty_data)
            col += 1
            cancel_reason_selection = [l[0] for l in line._fields['cancel_reason'].selection]
            worksheet.data_validation(row, 14, row, 14, {'validate': 'list', 'source': cancel_reason_selection})
            worksheet.write(row, col, line.cancel_reason or '', qty_data)
            col += 1
            row += 1

        workbook.close()
        return file_name

    def export_data_from_po_line(self):
        if self.env.context.get('active_model') == 'purchase.order':
            active_id = self.env.context.get("active_id")
            po_id = self.env["purchase.order"].browse(active_id)
            po_line_ids = self.env["purchase.order.line"]
            po_lines = po_id.order_line
            file_name = 'export_po_lines_data.xlsx'
            workbook = xlsxwriter.Workbook(f'/tmp/{file_name}')
            worksheet = workbook.add_worksheet("POLinesData")
            worksheet.protect()
            worksheet.set_landscape()
            worksheet.fit_to_pages(1, 0)
            worksheet.set_zoom(80)
            worksheet.set_column(1, 1, 20)
            worksheet.set_column(2, 2, 20)
            worksheet.set_column(3, 3, 15)
            worksheet.set_column(4, 4, 15)
            worksheet.set_column(6, 6, 20)

            qty_data = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'locked': False})
            qty_data_lock = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
            header_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'bold': True, 'border': 1})

            row, col = 0, 0
            worksheet.set_column(col, col, 30)
            worksheet.write(row, col, 'Isbn', header_format)
            col += 1
            worksheet.set_column(col, col, 40)
            worksheet.write(row, col, 'To Be Received QTY', header_format)
            col += 1
            worksheet.set_column(col, col, 15)
            worksheet.write(row, col, 'Bill List Price', header_format)
            col += 1
            worksheet.set_column(col, col, 20)
            worksheet.write(row, col, 'Bill Discount', header_format)
            col += 1
            row = 1

            records = [{rec.id: rec.product_id.id} for rec in po_lines]
            value_counts = defaultdict(int)
            for record in records:
                value = list(record.values())[0]
                value_counts[value] += 1

            unique_values = [record for record in records if value_counts[list(record.values())[0]] == 1]
            duplicate_values = [record for record in records if value_counts[list(record.values())[0]] > 1]

            total_to_be_received_qty = 0
            if duplicate_values:
                product_info_dict = {}
                for record in duplicate_values:
                    product_id = list(record.values())[0]
                    if product_id not in product_info_dict:
                        product_info_dict[product_id] = {
                            'barcode': '',
                            'total_to_be_received_qty': 0,
                            'list_price': 0,
                            'discount': 0,
                        }
                    product_info_dict[product_id]['barcode'] = po_line_ids.browse(record).product_id.barcode
                    product_info_dict[product_id]['total_to_be_received_qty'] += po_line_ids.browse(record).po_qty
                    product_info_dict[product_id]['list_price'] = po_line_ids.browse(record).list_price
                    product_info_dict[product_id]['discount'] = po_line_ids.browse(record).discount

                for product_id, info in product_info_dict.items():
                    col = 0
                    worksheet.write(row, col, info['barcode'], qty_data_lock)
                    col += 1
                    worksheet.write(row, col, info['total_to_be_received_qty'], qty_data)
                    col += 1
                    worksheet.write(row, col, info['list_price'], qty_data)
                    col += 1
                    worksheet.write(row, col, info['discount'], qty_data)
                    col += 1
                    row += 1

            if len(unique_values) > 0:
                all_keys = [key for dictionary in unique_values for key in dictionary.keys()]
                po_records = po_line_ids.browse(all_keys)
                for line in po_records:
                    col = 0
                    worksheet.write(row, col, line.product_id.barcode or '', qty_data_lock)
                    col += 1
                    worksheet.write(row, col, line.po_qty or 0, qty_data)
                    col += 1
                    worksheet.write(row, col, line.list_price or 0, qty_data)
                    col += 1
                    worksheet.write(row, col, line.discount or 0, qty_data)
                    col += 1
                    row += 1

            workbook.close()
            attachment = self.create_attachment(file_name)
            return {
                'type': 'ir.actions.act_url',
                'url': '/web/content/%s?download=true' % attachment.id,
                'target': 'download',
            }

    def import_po_line_from_export_file(self, field_list_to_import, sheet):
        purchase_id = self.env['purchase.order'].browse(self.env.context.get('active_id'))
        for row_no in range(sheet.nrows):
            if row_no != 0:
                passing_dict = dict(zip(field_list_to_import, sheet.row_values(row_no)))
                excel_list_price = passing_dict.get('bill list price')
                excel_discount = passing_dict.get('bill discount')
                if self.is_number(excel_list_price) and passing_dict.get('isbn'):
                    po_lines = purchase_id.order_line.filtered(lambda x: x.product_id.barcode == passing_dict.get('isbn'))
                    po_lines.update({'list_price' : float(excel_list_price)})
                if self.is_number(excel_discount) and passing_dict.get('isbn'):
                    po_lines = purchase_id.order_line.filtered(
                        lambda x: x.product_id.barcode == passing_dict.get('isbn'))
                    po_lines.update({'discount': float(excel_discount)})
                to_be_received_val = passing_dict.get('to be received qty')
                if to_be_received_val == '' or to_be_received_val is None:
                    continue
                to_be_received_qty = float(to_be_received_val) if self.is_number(to_be_received_val) else 0.0
                po_quantites = sum(purchase_id.order_line.filtered(lambda x: x.product_id.barcode == passing_dict.get('isbn')).mapped('po_qty'))
                bill_quantites = sum(purchase_id.order_line.filtered(lambda x: x.product_id.barcode == passing_dict.get('isbn')).mapped('qty_invoiced'))
                if po_quantites != bill_quantites:
                    if po_quantites < to_be_received_qty:
                        raise ValidationError(_('You can not set Vendor Bill Qty more then PO Quantity'))
                    elif po_quantites >= to_be_received_qty:
                        to_be_set_qty = to_be_received_qty
                        for rec in purchase_id.order_line.filtered(lambda x: x.product_id.barcode == passing_dict.get('isbn')):
                            if rec.po_qty == rec.qty_invoiced:
                                pass
                            else:
                                if to_be_set_qty >= rec.po_qty:
                                    rec.to_be_received_qty = rec.po_qty - rec.qty_invoiced if to_be_set_qty else 0
                                    to_be_set_qty = (to_be_set_qty - rec.to_be_received_qty) if to_be_set_qty else 0
                                elif to_be_set_qty < (rec.po_qty - rec.qty_invoiced):
                                    rec.to_be_received_qty = (to_be_set_qty) if to_be_set_qty > 0 else 0
                                    to_be_set_qty = to_be_set_qty - rec.to_be_received_qty
                                elif rec.qty_invoiced + to_be_set_qty > rec.po_qty:
                                    bill_qty_remain = rec.po_qty - rec.qty_invoiced
                                    if bill_qty_remain:
                                        rec.to_be_received_qty = bill_qty_remain
                                    to_be_set_qty = (to_be_set_qty - rec.to_be_received_qty) if to_be_set_qty else 0
                                else:
                                    rec.to_be_received_qty = to_be_set_qty
                                    to_be_set_qty -= to_be_set_qty
