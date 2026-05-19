# -*- coding: utf-8 -*-
import io
import base64
from datetime import date, timedelta
from odoo import api, fields, models, _
from odoo.exceptions import UserError

try:
    from odoo.tools.misc import xlsxwriter
except ImportError:
    import xlsxwriter

class PurchaseRegisterReportWizard(models.TransientModel):
    _name = 'purchase.register.report.wizard'
    _description = 'Purchase Register Report Wizard'

    date_range = fields.Selection([
        ('today', 'Today'),
        ('this_month', 'This Month'),
        ('this_quarter', 'This Quarter'),
        ('this_year', 'This Year'),
        ('last_month', 'Last Month'),
        ('last_quarter', 'Last Quarter'),
        ('last_year', 'Last Year'),
        ('2022', '2022'),
        ('2023', '2023'),
        ('2024', '2024'),
        ('2025', '2025'),
        ('2026', '2026'),
        ('custom', 'Custom')
    ], string='Date Range', default='custom')
    from_date = fields.Date(string='From Date', required=True)
    to_date = fields.Date(string='To Date', required=True)
    product_ids = fields.Many2many('product.product', string='Products')
    customer_ids = fields.Many2many('res.partner', string='Customers')
    company_ids = fields.Many2many('res.company', string='Companies')
    import_file = fields.Binary(string='Upload XLSX')
    import_file_name = fields.Char(string='File Name')

    def _get_products_from_xlsx(self):
        if not self.import_file:
            return self.env['product.product']

        try:
            from odoo.tools import pycompat
            import io
            import openpyxl
        except ImportError:
            raise UserError(_("The 'openpyxl' library is not installed."))

        try:
            # Decode the file
            file_content = base64.b64decode(self.import_file)
            file_stream = io.BytesIO(file_content)
            
            # Load workbook
            workbook = openpyxl.load_workbook(file_stream, data_only=True)
            sheet = workbook.active
            
            isbn_col_index = -1
            found_isbns = set()

            # Find 'ISBN' column in header
            for col_idx, cell in enumerate(sheet[1], 1): # Assuming header is in first row
                 if cell.value and str(cell.value).strip().upper() == 'ISBN':
                     isbn_col_index = col_idx - 1 # 0-based index
                     break
            
            if isbn_col_index == -1:
                 raise UserError(_("Column 'ISBN' not found in the uploaded XLSX file."))

            # Iterate rows to collect ISBNs
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[isbn_col_index] is not None:
                    isbn_val = row[isbn_col_index]
                    # Handle numbers that might be read as float/int
                    if isinstance(isbn_val, float):
                         if isbn_val.is_integer():
                             isbn_val = int(isbn_val)
                    isbn = str(isbn_val).strip()
                    if isbn:
                        found_isbns.add(isbn)

            if not found_isbns:
                raise UserError(_("No ISBNs found in the 'ISBN' column."))

            # Find products - Case insensitive search for better matching
            # Also handle if ISBNs are stored as slightly different strings in Odoo
            domain = ['|', ('default_code', 'in', list(found_isbns)), ('barcode', 'in', list(found_isbns))]
            
            # If standard search fails, try ILIKE for single values if list is small, 
            # but for bulk 'in' operator is best. We stick to exact match for performance 
            # but rely on the cleaned list.
            
            products = self.env['product.product'].search(domain)
            
            if not products:
                 # Debug info for user
                 fail_msg = _("No products found for the provided ISBNs.\nScanned %d unique ISBNs from file.") % len(found_isbns)
                 fail_msg += "\n" + _("Example ISBNs from file: %s") % (', '.join(list(found_isbns)[:5]))
                 raise UserError(fail_msg)

            return products

        except Exception as e:
            raise UserError(_("Error reading file: %s") % str(e))

    def action_import_xlsx(self):
        if not self.import_file:
            raise UserError(_("Please upload an XLSX file."))
        
        products = self._get_products_from_xlsx()
        self.product_ids = [(6, 0, products.ids)]

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'purchase.register.report.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

    @api.onchange('date_range')
    def _onchange_date_range(self):
        today = date.today()
        if self.date_range == 'today':
            self.from_date = today
            self.to_date = today
        elif self.date_range == 'this_month':
            self.from_date = today.replace(day=1)
            import calendar
            last_day = calendar.monthrange(today.year, today.month)[1]
            self.to_date = today.replace(day=last_day)
        elif self.date_range == 'last_month':
            first_day_this_month = today.replace(day=1)
            last_day_last_month = first_day_this_month - timedelta(days=1)
            self.from_date = last_day_last_month.replace(day=1)
            self.to_date = last_day_last_month
        elif self.date_range == 'this_quarter':
            quarter = (today.month - 1) // 3 + 1
            self.from_date = today.replace(month=(quarter - 1) * 3 + 1, day=1)
            last_month = quarter * 3
            import calendar
            last_day = calendar.monthrange(today.year, last_month)[1]
            self.to_date = today.replace(month=last_month, day=last_day)
        elif self.date_range == 'last_quarter':
            quarter = (today.month - 1) // 3 + 1
            if quarter == 1:
                year = today.year - 1
                last_quarter = 4
            else:
                year = today.year
                last_quarter = quarter - 1
            self.from_date = today.replace(year=year, month=(last_quarter - 1) * 3 + 1, day=1)
            last_month = last_quarter * 3
            import calendar
            last_day = calendar.monthrange(year, last_month)[1]
            self.to_date = today.replace(year=year, month=last_month, day=last_day)
        elif self.date_range == 'this_year':
            self.from_date = today.replace(month=1, day=1)
            self.to_date = today.replace(month=12, day=31)
        elif self.date_range == 'last_year':
            self.from_date = today.replace(year=today.year - 1, month=1, day=1)
            self.to_date = today.replace(year=today.year - 1, month=12, day=31)
        elif self.date_range in ['2022', '2023', '2024', '2025', '2026']:
            year = int(self.date_range)
            self.from_date = date(year, 1, 1)
            self.to_date = date(year, 12, 31)

    def action_export_xlsx(self):
        if self.from_date > self.to_date:
            raise UserError(_("From Date must be less than To Date."))

        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        sheet = workbook.add_worksheet("Purchase Register Report")

        # Define Formats
        header_format = workbook.add_format({
            'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#D3D3D3'
        })
        cell_format = workbook.add_format({'align': 'left', 'valign': 'vcenter', 'border': 1})
        right_format = workbook.add_format({'align': 'right', 'valign': 'vcenter', 'border': 1})
        center_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})
        date_format = workbook.add_format({'num_format': 'yyyy-mm-dd', 'align': 'left', 'valign': 'vcenter', 'border': 1})
        price_format = workbook.add_format({'num_format': '#,##0.00', 'align': 'right', 'valign': 'vcenter', 'border': 1})
        group_header_format = workbook.add_format({
            'bold': True, 'bg_color': '#E9ECEF', 'border': 1, 'valign': 'vcenter'
        })
        subtotal_format = workbook.add_format({
            'bold': True, 'border': 1, 'bg_color': '#F8F9FA', 'align': 'right'
        })

        # Set Column Widths
        sheet.set_column(0, 0, 18)  # ISBN
        sheet.set_column(1, 1, 30)  # Product Name
        sheet.set_column(2, 2, 12)  # Bill Date
        sheet.set_column(3, 3, 20)  # Bill Number
        sheet.set_column(4, 4, 20)  # Bill Reference
        sheet.set_column(5, 5, 25)  # Vendor Name
        sheet.set_column(6, 6, 20)  # Company
        sheet.set_column(7, 7, 12)  # Currency
        sheet.set_column(8, 8, 20)  # CSO
        sheet.set_column(9, 9, 25)  # Customer Name
        sheet.set_column(10, 10, 10)  # Quantity
        sheet.set_column(11, 11, 12) # List Price
        sheet.set_column(12, 12, 12) # Discount (%)
        sheet.set_column(13, 13, 12) # Unit Price
        sheet.set_column(14, 14, 12) # Subtotal
        sheet.set_column(15, 15, 12) # Total Amount

        # Headers
        headers = [
            'ISBN', 'Product Name', 'Bill Date', 'Bill Number', 'Bill Reference',
            'Vendor Name', 'Company', 'Currency', 'CSO', 'Customer Name', 'Quantity', 'List Price',
            'Discount (%)', 'Unit Price', 'Subtotal', 'Total Amount'
        ]
        for col_num, header in enumerate(headers):
            sheet.write(0, col_num, header, header_format)

        # Build SQL Query
        query = """
            SELECT
                prod.default_code as isbn,
                tmpl.name as product_name,
                move.invoice_date as bill_date,
                move.name as bill_number,
                move.ref as bill_reference,
                rp.name as vendor_name,
                company.name as company_name,
                curr.name as currency,
                COALESCE(line.customer_sales_order, move.customer_sales_order) as cso,
                COALESCE(line.customer_name, move.customer_name) as customer_name,
                line.quantity as quantity,
                line.list_price as list_price,
                line.discount as discount,
                line.price_unit as unit_price,
                line.price_subtotal as subtotal,
                line.price_total as total_amount,
                prod.id as product_id
            FROM
                account_move_line line
            JOIN
                account_move move ON (line.move_id = move.id)
            JOIN
                res_company company ON (move.company_id = company.id)
            JOIN
                res_currency curr ON (move.currency_id = curr.id)
            JOIN
                product_product prod ON (line.product_id = prod.id)
            JOIN
                product_template tmpl ON (prod.product_tmpl_id = tmpl.id)
            JOIN
                res_partner rp ON (move.partner_id = rp.id)
            WHERE
                move.move_type = 'in_invoice'
                AND move.state = 'posted'
                AND move.invoice_date >= %s
                AND move.invoice_date <= %s
                AND line.display_type NOT IN ('line_section', 'line_note')
                AND tmpl.type != 'service'
                AND line.quantity > 0
        """
        params = [self.from_date, self.to_date]

        if self.customer_ids:
            query += " AND move.partner_id IN %s"
            params.append(tuple(self.customer_ids.ids))

        if self.company_ids:
            query += " AND move.company_id IN %s"
            params.append(tuple(self.company_ids.ids))

        # Combine manually selected products and products from XLSX
        filter_product_ids = set(self.product_ids.ids)
        if self.import_file:
             xlsx_products = self._get_products_from_xlsx()
             filter_product_ids.update(xlsx_products.ids)

        if filter_product_ids:
            query += " AND prod.id IN %s"
            params.append(tuple(filter_product_ids))

        # Order by product to facilitate grouping
        query += " ORDER BY tmpl.name, prod.id, move.invoice_date"

        self.env.cr.execute(query, params)
        results = self.env.cr.dictfetchall()

        # Handle translatable fields (JSONB columns return dict in Odoo 17+)
        lang = self.env.lang or 'en_US'
        for res in results:
            for key in res:
                if isinstance(res[key], dict):
                    res[key] = res[key].get(lang) or res[key].get('en_US') or (list(res[key].values())[0] if res[key] else '')

        row = 1
        current_product_id = None
        group_qty = 0
        group_subtotal = 0
        group_total = 0

        for res in results:
            if current_product_id != res['product_id']:
                # If it's not the first product, write the subtotal for the previous group
                if current_product_id is not None:
                    sheet.write(row, 9, "Subtotal:", subtotal_format)
                    sheet.write(row, 10, group_qty, subtotal_format)
                    sheet.write(row, 14, group_subtotal, price_format)
                    sheet.write(row, 15, group_total, price_format)
                    row += 1

                # Reset group counters and write group header
                current_product_id = res['product_id']
                group_qty = 0
                group_subtotal = 0
                group_total = 0
                
                sheet.merge_range(row, 0, row, 15, f"Product: [{res['isbn'] or 'N/A'}] {res['product_name']}", group_header_format)
                row += 1

            # Write Data Row
            sheet.write(row, 0, res['isbn'] or '', cell_format)
            sheet.write(row, 1, res['product_name'] or '', cell_format)
            sheet.write(row, 2, res['bill_date'] or '', date_format)
            sheet.write(row, 3, res['bill_number'] or '', cell_format)
            sheet.write(row, 4, res['bill_reference'] or '', cell_format)
            sheet.write(row, 5, res['vendor_name'] or '', cell_format)
            sheet.write(row, 6, res['company_name'] or '', cell_format)
            sheet.write(row, 7, res['currency'] or '', center_format)
            sheet.write(row, 8, res['cso'] or '', cell_format)
            sheet.write(row, 9, res['customer_name'] or '', cell_format)
            sheet.write(row, 10, res['quantity'], right_format)
            sheet.write(row, 11, res['list_price'], price_format)
            sheet.write(row, 12, res['discount'], right_format)
            sheet.write(row, 13, res['unit_price'], price_format)
            sheet.write(row, 14, res['subtotal'], price_format)
            sheet.write(row, 15, res['total_amount'], price_format)

            group_qty += res['quantity']
            group_subtotal += res['subtotal']
            group_total += res['total_amount']
            row += 1

        # Write the subtotal for the final product group
        if current_product_id is not None:
            sheet.write(row, 9, "Subtotal:", subtotal_format)
            sheet.write(row, 10, group_qty, subtotal_format)
            sheet.write(row, 14, group_subtotal, price_format)
            sheet.write(row, 15, group_total, price_format)
            row += 1

        workbook.close()
        output.seek(0)
        file_data = base64.b64encode(output.getvalue())
        output.close()

        attachment = self.env['ir.attachment'].create({
            'name': 'Purchase_Register_Report.xlsx',
            'datas': file_data,
            'res_model': 'purchase.register.report.wizard',
            'res_id': self.id,
            'type': 'binary',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }
